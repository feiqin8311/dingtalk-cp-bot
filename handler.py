#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""DingTalk message handler for querying LingXing shipments."""

from __future__ import annotations

import asyncio
import base64
import json
import logging
import random
import re
import shutil
import sys
import threading
import time
import traceback
import urllib.error
import urllib.request
import uuid
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Awaitable, Callable, Dict, Iterable, List, Optional, Tuple
import os

import dingtalk_stream
import pymysql
from dingtalk_stream import AckMessage
import fitz  # PyMuPDF

import config


def _bootstrap_common_import_path() -> List[Path]:
    candidates: List[Path] = []
    configured = os.getenv("COMMON_DIR", "").strip()
    if configured:
        candidates.append(Path(configured).expanduser())
    candidates.append(config.COMMON_DIR)
    base_dir = Path(__file__).resolve().parent
    for parent in (base_dir, *base_dir.parents):
        candidates.append(parent / "Common")

    valid_common_dirs: List[Path] = []
    seen = set()
    for raw in candidates:
        try:
            path = raw.resolve()
        except Exception:
            path = raw
        key = str(path).lower()
        if key in seen:
            continue
        seen.add(key)

        common_dir = None
        if path.exists() and path.is_dir() and (path / "api").exists():
            common_dir = path
        elif path.exists() and path.is_dir() and path.name.lower() == "api":
            parent_dir = path.parent
            if (parent_dir / "api").exists():
                common_dir = parent_dir

        if not common_dir:
            continue
        common_text = str(common_dir)
        if common_text not in sys.path:
            sys.path.insert(0, common_text)
        valid_common_dirs.append(common_dir)

    return valid_common_dirs


_COMMON_DIRS = _bootstrap_common_import_path()

from api import DingTalkNotifier  # type: ignore
from api.aliyun_client import AliyunOCRClient  # type: ignore
from api.lingxing_client import LingXingClient  # type: ignore


_DEFAULT_NOTIFIER = DingTalkNotifier(
    app_key=config.DINGTALK_APP_KEY,
    app_secret=config.DINGTALK_APP_SECRET,
    robot_code=config.DINGTALK_ROBOT_CODE,
)


def _post_json(url: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, method="POST")
    req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            body = resp.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"request failed: {exc.code} {exc.reason}. {detail}") from exc
    if not body:
        return {}
    return json.loads(body)


def _send_session_text(session_webhook: str, text: str) -> Dict[str, Any]:
    return _post_json(session_webhook, {"msgtype": "text", "text": {"content": text}})


def _unique_preserve_order(items: Iterable[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result


def _log_json(logger: logging.Logger, label: str, payload: Any, request_id: Optional[str] = None) -> None:
    try:
        text = json.dumps(payload, ensure_ascii=False)
    except Exception:
        text = str(payload)
    prefix = f"[req={request_id}] " if request_id else ""
    logger.info("%s%s: %s", prefix, label, text)


def _summarize_lingxing_response(response: Dict[str, Any]) -> Dict[str, Any]:
    code = response.get("code") or response.get("Code") or response.get("status")
    msg = response.get("message") or response.get("msg") or response.get("Message") or ""
    data = response.get("data")
    data_list = ShipmentQueryHandler._normalize_data_list(data)
    shipments: List[Dict[str, Any]] = []
    for shipment in data_list:
        shipment_sn = (
            shipment.get("shipment_sn")
            or shipment.get("shipment_no")
            or shipment.get("shipmentNo")
            or ""
        )
        items = shipment.get("items") or []
        nation = _extract_first_item_nation(items)
        dest_ids = []
        if isinstance(items, list):
            for item in items:
                if not isinstance(item, dict):
                    continue
                dest = item.get("destination_fulfillment_center_id")
                if dest:
                    dest_ids.append(str(dest).strip().upper())
        dest_ids = _unique_preserve_order([d for d in dest_ids if d])

        file_list = shipment.get("fileList") or shipment.get("file_list") or []
        pdf_files = []
        if isinstance(file_list, list):
            for file_info in file_list:
                if not isinstance(file_info, dict):
                    continue
                name = str(
                    file_info.get("file_name")
                    or file_info.get("fileName")
                    or ""
                ).strip()
                file_id = file_info.get("file_id") or file_info.get("fileId")
                if name.upper().startswith("FBA") and name.lower().endswith(".pdf"):
                    pdf_files.append(
                        {
                            "file_name": name,
                            "file_id": file_id,
                        }
                    )

        shipments.append(
            {
                "shipment_sn": shipment_sn,
                "nation": nation,
                "destinations": dest_ids,
                "pdf_files": pdf_files,
                "file_count": len(file_list) if isinstance(file_list, list) else 0,
            }
        )

    return {
        "code": code,
        "message": msg,
        "data_count": len(data_list),
        "shipments": shipments,
    }


def _download_file(url: str, dest_path: Path, *, expect_pdf: bool = False) -> Path:
    req = urllib.request.Request(
        url,
        method="GET",
        headers={"User-Agent": "Mozilla/5.0"},
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        content_type = resp.headers.get("Content-Type", "")
        data = resp.read()
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    dest_path.write_bytes(data)

    if expect_pdf and not data.startswith(b"%PDF"):
        header = data[:8]
        raise RuntimeError(
            f"downloaded file is not PDF (content-type={content_type}, size={len(data)}, header={header.hex()})"
        )
    return dest_path


def _save_pdf_bytes(data: bytes, dest_path: Path) -> Path:
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    dest_path.write_bytes(data)
    if not data.startswith(b"%PDF"):
        header = data[:8]
        raise RuntimeError(f"downloaded file is not PDF (size={len(data)}, header={header.hex()})")
    return dest_path


def _pdf_first_page_to_png(pdf_path: Path, image_path: Path) -> Path:
    doc = fitz.open(str(pdf_path))
    try:
        page = doc.load_page(0)
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        image_path.parent.mkdir(parents=True, exist_ok=True)
        pix.save(str(image_path))
    finally:
        doc.close()
    return image_path


def _extract_shipment_sns(text: str) -> List[str]:
    if not text:
        return []
    matches = re.findall(r"\bSP[0-9A-Za-z]+\b", text, flags=re.IGNORECASE)
    if matches:
        return _unique_preserve_order([m.strip().upper() for m in matches])
    normalized = re.sub(r"[,\s，、;；]+", " ", text.strip())
    tokens = [tok.strip() for tok in normalized.split(" ") if tok.strip()]
    return _unique_preserve_order(tokens)


def _parse_file_selection_command(text: str) -> Optional[Tuple[str, int]]:
    raw = str(text or "").strip()
    if not raw:
        return None
    match = re.match(
        r"^\s*(?:选择文件|选文件|pick)\s+(SP[0-9A-Za-z]+)\s+(\d+)\s*$",
        raw,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    shipment_sn = str(match.group(1)).strip().upper()
    index = int(match.group(2))
    if index <= 0:
        return None
    return shipment_sn, index


def _extract_text_content(incoming_message) -> Optional[str]:
    text = None
    if hasattr(incoming_message, "text") and incoming_message.text is not None:
        text_obj = incoming_message.text
        if isinstance(text_obj, str):
            text = text_obj
        elif isinstance(text_obj, dict):
            text = text_obj.get("content") or text_obj.get("text")
        else:
            text = getattr(text_obj, "content", None) or getattr(text_obj, "text", None)
    if not text and hasattr(incoming_message, "content"):
        text = getattr(incoming_message, "content", None)
    if not text and hasattr(incoming_message, "text_content"):
        text = getattr(incoming_message, "text_content", None)
    return str(text).strip() if text else None


def _extract_session_webhook(incoming_message, raw_data) -> Optional[str]:
    for attr in ("session_webhook", "sessionWebhook"):
        if hasattr(incoming_message, attr):
            value = getattr(incoming_message, attr)
            if value:
                return value
    if isinstance(raw_data, str):
        try:
            raw_data = json.loads(raw_data)
        except Exception:
            raw_data = None
    if isinstance(raw_data, dict):
        for key in ("sessionWebhook", "session_webhook"):
            value = raw_data.get(key)
            if value:
                return value
        nested = raw_data.get("data")
        if isinstance(nested, dict):
            for key in ("sessionWebhook", "session_webhook"):
                value = nested.get(key)
                if value:
                    return value
    return None


def _extract_sender_user_id(incoming_message) -> Optional[str]:
    for attr in (
        "sender_staff_id",
        "senderStaffId",
        "sender_id",
        "senderId",
        "sender_open_id",
        "senderOpenId",
    ):
        if hasattr(incoming_message, attr):
            value = getattr(incoming_message, attr)
            if value:
                return str(value)
    return None


def _extract_sender_user_name(incoming_message, raw_data: Any = None) -> Optional[str]:
    for attr in (
        "sender_nick",
        "senderNick",
        "sender_name",
        "senderName",
        "sender_staff_name",
        "senderStaffName",
    ):
        if hasattr(incoming_message, attr):
            value = getattr(incoming_message, attr)
            if value:
                return str(value).strip()
    if isinstance(raw_data, str):
        try:
            raw_data = json.loads(raw_data)
        except Exception:
            raw_data = None
    if isinstance(raw_data, dict):
        for key in ("senderNick", "sender_nick", "senderName", "sender_name", "senderStaffName"):
            value = raw_data.get(key)
            if value:
                return str(value).strip()
        nested = raw_data.get("data")
        if isinstance(nested, dict):
            for key in ("senderNick", "sender_nick", "senderName", "sender_name", "senderStaffName"):
                value = nested.get(key)
                if value:
                    return str(value).strip()
    return None


def _extract_ocr_text(result: Dict[str, Any]) -> str:
    data = result.get("data") or result.get("Data") or result.get("body") or result
    if isinstance(data, dict):
        content = data.get("content") or data.get("Content")
        if content:
            return str(content)
        for key in ("prism_wordsInfo", "prism_words_info", "wordsInfo", "words_info"):
            words_info = data.get(key)
            if isinstance(words_info, list):
                words = []
                for item in words_info:
                    if not isinstance(item, dict):
                        continue
                    word = item.get("word") or item.get("words") or item.get("text") or item.get("content")
                    if word:
                        words.append(str(word))
                if words:
                    return "\n".join(words)
    return ""


def _extract_fc_code(text: str) -> Optional[str]:
    if not text:
        return None
    upper = text.upper()
    patterns = [
        r"FBA[:\s]*([A-Z]{3}\d)",
        r"目的地[:\s]*([A-Z]{3}\d)",
        r"\b([A-Z]{3}\d)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, upper)
        if match:
            return match.group(1)
    return None


def _extract_fc_codes(text: str) -> List[str]:
    if not text:
        return []
    upper = str(text).upper()
    matches = re.findall(r"\b([A-Z]{3}\d)\b", upper)
    return _unique_preserve_order([str(item).strip().upper() for item in matches if str(item).strip()])


def _extract_fc_codes_filtered(text: str) -> List[str]:
    if not text:
        return []
    upper = str(text).upper()
    results: List[str] = []

    def _allow_by_context(start: int, end: int) -> bool:
        left = max(0, start - 28)
        right = min(len(upper), end + 28)
        window = upper[left:right]
        if "SINGLE SKU" in window:
            return False
        if "数量" in window or "QTY" in window:
            return False
        if re.search(r"\bPM\d{4,}\b", window):
            return False
        return True

    # Capture pair forms first, e.g. "LBA8/IMN1" (including connected text like ServicesLBA8/IMN1).
    for match in re.finditer(r"([A-Z]{3}\d)\s*[\\/／]\s*([A-Z]{3}\d)", upper):
        if not _allow_by_context(match.start(), match.end()):
            continue
        results.append(str(match.group(1)).strip().upper())
        results.append(str(match.group(2)).strip().upper())

    # Fallback single code scan (do not require word boundary, only avoid digit-continuations).
    for match in re.finditer(r"(?<!\d)([A-Z]{3}\d)(?!\d)", upper):
        code = str(match.group(1)).strip().upper()
        if not _allow_by_context(match.start(1), match.end(1)):
            continue
        results.append(code)

    return _unique_preserve_order(results)


def _extract_first_item_nation(items: Any) -> Optional[str]:
    if not isinstance(items, list):
        return None
    for item in items:
        if not isinstance(item, dict):
            continue
        nation = item.get("nation")
        if nation is None:
            return None
        nation_text = str(nation).strip()
        return nation_text or None
    return None


ADDRESS_MATCH_PASS_SCORE = 82.0
ADDRESS_MATCH_MIN_GAP = 8.0
SUPPORTED_ADDRESS_COUNTRIES = {"US", "DE", "UK", "AU"}

_ROAD_KEYWORDS = (
    "ROAD",
    "STREET",
    "DRIVE",
    "AVENUE",
    "PASEO",
    "LANE",
    "WAY",
    "BOULEVARD",
    "BLVD",
    "COURT",
    "PLACE",
    "TERRACE",
    "HIGHWAY",
    "STRASSE",
)
_STREET_ABBR_WORD_MAP = {
    "RD": "ROAD",
    "ST": "STREET",
    "DR": "DRIVE",
    "AVE": "AVENUE",
    "BLVD": "BOULEVARD",
    "STR": "STRASSE",
}
_COMPANY_SUFFIX_TOKENS = {
    "INC",
    "INCORPORATED",
    "LLC",
    "LTD",
    "LIMITED",
    "CO",
    "COMPANY",
    "GMBH",
    "SARL",
}
_COUNTRY_WORD_TOKENS = {
    "US",
    "USA",
    "UNITED",
    "STATES",
    "GERMANY",
    "DEUTSCHLAND",
    "UK",
    "ENGLAND",
    "AU",
    "AUSTRALIA",
}

_UK_POSTAL_RE = re.compile(
    r"([A-Z]{1,2}\d[A-Z\d]?\s?\d[A-Z]{2})(?=(?:\b|ENGLAND\b|UK\b|UNITED\s*KINGDOM\b|$))"
)
_US_CITY_STATE_ZIP_RE = re.compile(
    r"\b([A-Z][A-Z\-]*(?:\s+[A-Z][A-Z\-]*){0,3})\s+([A-Z]{2})\s+(\d{5}(?:\d{4})?)\b"
)
_US_STATE_ZIP_RE = re.compile(r"\b([A-Z]{2})\s+(\d{5}(?:\d{4})?)\b")
_DE_ZIP_CITY_RE = re.compile(r"\b(\d{5})\s*([A-Z][A-Z\s\-]+)\b")
_AU_CITY_STATE_POSTAL_RE = re.compile(r"\b([A-Z][A-Z\s\-]+)\s+(NSW|VIC|QLD|WA|SA|TAS|ACT|NT)\s+(\d{4})\b")
_NOISE_LINE_PREFIX_RE = re.compile(
    r"^\s*(ANMELDER|DECLARANT|CONSIGNEE|ATTN|ATTENTION|收件人)\s*[:：]",
    re.IGNORECASE,
)
_COUNTRY_LINE_RE = re.compile(
    r"^(美国|德国|英国|澳大利亚|澳洲|US|USA|GERMANY|DEUTSCHLAND|UK|ENGLAND|AUSTRALIA|AU)$",
    re.IGNORECASE,
)
_SOURCE_NOISE_TOKENS = {
    "ZHEJIANG",
    "HANGZHOU",
    "XIHU",
    "CHINA",
    "ROOM",
    "BUILDING",
    "DISTRICT",
    "BUSINESS",
    "CENTER",
    "WENYIXI",
    "WENRYIXI",
    "JIANGCUN",
    "YPLUSEU",
    "LIBRATON",
}
_DE_CITY_STOP_TOKENS = _SOURCE_NOISE_TOKENS | {"ROAD", "STRASSE", "NO"}
_UK_REGION_TOKENS = {"WEST", "EAST", "NORTH", "SOUTH", "MIDLANDS", "GREATER"}
_UK_STREET_DROP_TOKENS = {
    "BU",
    "B",
    "ROOM",
    "NO",
    "ENTER",
    "CENTER",
    "C",
    "ILDINGB",
    "BUILDINGB",
    "EU",
    "SARL",
    "UK",
    "AMAZON",
}


@dataclass
class _AddressRecord:
    country_mode: str
    raw_text: str
    fc_code: str
    company: str = ""
    street: str = ""
    street_no: str = ""
    city: str = ""
    state: str = ""
    zip5: str = ""
    postal_code: str = ""


@dataclass
class _AddressCandidateScore:
    record: _AddressRecord
    score: float
    hard_ok: bool
    hard_reason: str


@dataclass
class _AddressMatchOutcome:
    fc_code: str
    score: float
    gap: float
    postal_code: str
    city: str
    ocr_extracted: str
    excel_extracted: str


def _normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _normalize_text_for_match(text: str) -> str:
    normalized = str(text or "").upper()
    normalized = (
        normalized.replace("Ä", "AE")
        .replace("Ö", "OE")
        .replace("Ü", "UE")
        .replace("ß", "SS")
    )
    normalized = re.sub(r"[^A-Z0-9\s]", " ", normalized)
    return _normalize_spaces(normalized)


def _normalize_uk_postal_code(postal: str) -> str:
    compact = re.sub(r"[^A-Z0-9]", "", str(postal or "").upper())
    if len(compact) < 5:
        return ""
    return f"{compact[:-3]} {compact[-3:]}"


def _normalize_au_state(value: str) -> str:
    text = _normalize_text_for_match(value)
    text = text.replace("N S W", "NSW").replace("N.S.W", "NSW")
    return text


def _normalize_fc_code(value: str) -> str:
    compact = re.sub(r"[^A-Z0-9]", "", str(value or "").upper())
    if not compact:
        return ""
    match = re.search(r"[A-Z]{3}\d", compact)
    if match:
        return match.group(0)
    return compact


def _detect_country_mode_by_nation(nation: Optional[str]) -> Optional[str]:
    raw = str(nation or "").strip()
    if not raw:
        return None
    upper = raw.upper()
    if "美国" in raw or "UNITED STATES" in upper or re.search(r"\bUSA?\b", upper):
        return "US"
    if "德国" in raw or "GERMANY" in upper or "DEUTSCHLAND" in upper:
        return "DE"
    if "英国" in raw or "UNITED KINGDOM" in upper or "ENGLAND" in upper or re.search(r"\bUK\b", upper):
        return "UK"
    if "澳洲" in raw or "澳大利亚" in raw or "AUSTRALIA" in upper or re.search(r"\bAU\b", upper):
        return "AU"
    return None


def _is_fc_contains_country(nation: Optional[str]) -> bool:
    text = str(nation or "").strip()
    if not text:
        return False
    return any(country in text for country in ("加拿大", "日本", "阿联酋"))


def _dest_code_contains_in_ocr_text(dest_ids: List[str], ocr_text: str) -> bool:
    compact_ocr = re.sub(r"[^A-Z0-9]+", "", str(ocr_text or "").upper())
    if not compact_ocr:
        return False
    for dest_id in _unique_preserve_order(dest_ids):
        code = _normalize_fc_code(dest_id)
        if not code:
            continue
        if code in compact_ocr:
            return True
    return False


def _extract_destination_block(text: str) -> str:
    one_line = _normalize_spaces(str(text or "").replace("\r", " ").replace("\n", " "))
    if one_line:
        same_line_match = re.search(
            r"(?:目的地|DESTINATION)\s*[:：]?\s*(.*?)\s*(?:发货地|SHIP\s*FROM|SHIPFROM)\s*[:：]?",
            one_line,
            flags=re.IGNORECASE,
        )
        if same_line_match:
            return _normalize_spaces(same_line_match.group(1))

    lines = [line.strip() for line in str(text or "").replace("\r", "\n").splitlines() if line.strip()]
    if not lines:
        return ""

    start_idx: Optional[int] = None
    prefix_line: Optional[str] = None
    for index, line in enumerate(lines):
        if re.search(r"(目的地|DESTINATION)\s*[:：]?", line, flags=re.IGNORECASE):
            start_idx = index + 1
            line_after = re.split(r"[:：]", line, maxsplit=1)
            if len(line_after) == 2:
                prefix_line = line_after[1].strip()
            break

    if start_idx is None:
        return "\n".join(lines)

    end_idx = len(lines)
    for index in range(start_idx, len(lines)):
        if re.search(r"(发货地|SHIP\s*FROM|SHIPFROM)\s*[:：]?", lines[index], flags=re.IGNORECASE):
            end_idx = index
            break

    selected = []
    if prefix_line:
        selected.append(prefix_line)
    selected.extend(lines[start_idx:end_idx])
    return "\n".join([item for item in selected if item])


def _normalize_street_text(text: str, country_mode: str) -> str:
    normalized = _normalize_text_for_match(text)
    normalized = re.sub(r"(?<=\d)(?=[A-Z])", " ", normalized)
    normalized = re.sub(r"(?<=[A-Z])(?=\d)", " ", normalized)
    normalized = re.sub(r"\b(ROAD)(ROOM)\b", r"\1 \2", normalized)
    normalized = re.sub(r"\b(STREET)(ROOM)\b", r"\1 \2", normalized)
    normalized = re.sub(r"\b(DRIVE)(ROOM)\b", r"\1 \2", normalized)
    normalized = re.sub(r"\b(LANE)(ROOM)\b", r"\1 \2", normalized)
    normalized = re.sub(r"\b(AVENUE)(ROOM)\b", r"\1 \2", normalized)
    normalized = re.sub(r"\b([A-Z]{3,})(DR|RD|ST|AVE)\b", r"\1 \2", normalized)
    for short, full in _STREET_ABBR_WORD_MAP.items():
        normalized = re.sub(rf"\b{short}\b", full, normalized)
    if country_mode == "DE":
        normalized = re.sub(r"\bSTRASSE\b", "STRASSE", normalized)
    return normalized


def _extract_street_no(street: str, country_mode: str) -> str:
    text = _normalize_text_for_match(street)
    if country_mode == "DE":
        match = re.search(r"(\d+[A-Z]?)$", text)
        if match:
            return match.group(1)
    # For US/UK/AU, skip long numeric chunks (e.g. 310030 from source-side postal noise).
    for match in re.finditer(r"\b(\d+[A-Z]?)\b", text):
        candidate = match.group(1)
        digits = re.sub(r"\D", "", candidate)
        if not digits:
            continue
        if len(digits) <= 5:
            return candidate
    return ""


def _street_similarity(a: str, b: str) -> float:
    ta = _normalize_street_text(a, "US")
    tb = _normalize_street_text(b, "US")
    if not ta or not tb:
        return 0.0
    tokens_a = set(ta.split())
    tokens_b = set(tb.split())
    jaccard = (len(tokens_a & tokens_b) / len(tokens_a | tokens_b)) if (tokens_a or tokens_b) else 0.0
    seq = SequenceMatcher(None, ta, tb).ratio()
    return 0.6 * seq + 0.4 * jaccard


def _uk_street_token_coverage(source_street: str, candidate_street: str) -> float:
    source = _normalize_street_text(source_street, "UK")
    candidate = _normalize_street_text(candidate_street, "UK")
    if not source or not candidate:
        return 0.0

    def clean_tokens(text: str) -> List[str]:
        tokens: List[str] = []
        for token in text.split():
            if token in _UK_STREET_DROP_TOKENS or token in _SOURCE_NOISE_TOKENS:
                continue
            if token in _COUNTRY_WORD_TOKENS:
                continue
            if re.fullmatch(r"\d{3,}", token):
                continue
            tokens.append(token)
        return tokens

    source_set = set(clean_tokens(source))
    candidate_tokens = clean_tokens(candidate)
    if not source_set or not candidate_tokens:
        return 0.0
    required = [token for token in candidate_tokens if token not in {"WEST", "MIDLANDS"}]
    if not required:
        required = candidate_tokens
    matched = sum(1 for token in required if token in source_set)
    return matched / len(required)


def _company_similarity(a: str, b: str) -> float:
    ta = [token for token in _normalize_text_for_match(a).split() if token not in _COMPANY_SUFFIX_TOKENS]
    tb = [token for token in _normalize_text_for_match(b).split() if token not in _COMPANY_SUFFIX_TOKENS]
    if not ta or not tb:
        return 0.0
    a_text = " ".join(ta)
    b_text = " ".join(tb)
    return SequenceMatcher(None, a_text, b_text).ratio()


def _city_matches(expected: str, actual: str) -> bool:
    left = _normalize_text_for_match(expected)
    right = _normalize_text_for_match(actual)
    if not left or not right:
        return False
    if left == right:
        return True
    if left in right or right in left:
        return True

    left_tokens = [t for t in left.split() if t and t not in _ROAD_KEYWORDS and not re.search(r"\d", t)]
    right_tokens = [t for t in right.split() if t and t not in _ROAD_KEYWORDS and not re.search(r"\d", t)]
    if not left_tokens or not right_tokens:
        return False
    if left_tokens[-1] == right_tokens[-1] and len(left_tokens[-1]) >= 4:
        return True
    overlap = set(left_tokens) & set(right_tokens)
    if overlap and len(overlap) >= min(len(left_tokens), len(right_tokens)):
        return True
    # OCR city tokens may contain minor character errors (e.g. HAZLE vs HAZIE).
    # Accept when normalized whole-city similarity is sufficiently high.
    if SequenceMatcher(None, " ".join(left_tokens), " ".join(right_tokens)).ratio() >= 0.72:
        return True
    return False


def _is_country_line(line: str) -> bool:
    normalized = _normalize_text_for_match(line)
    if not normalized:
        return True
    if _COUNTRY_LINE_RE.match(line.strip()):
        return True
    if normalized in {"US", "USA", "GERMANY", "UK", "ENGLAND", "AUSTRALIA", "AU"}:
        return True
    return False


def _is_source_noise_line(line: str) -> bool:
    normalized = _normalize_text_for_match(line)
    if not normalized:
        return False
    tokens = set(normalized.split())
    return any(token in tokens for token in _SOURCE_NOISE_TOKENS)


def _prepare_address_lines(text: str) -> List[str]:
    lines: List[str] = []
    for raw_line in str(text or "").replace("\r", "\n").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^\s*(目的地|DESTINATION)\s*[:：]\s*", "", line, flags=re.IGNORECASE)
        line = re.sub(r"^\s*(发货地|SHIP\s*FROM|SHIPFROM)\s*[:：]\s*", "", line, flags=re.IGNORECASE)
        line = line.strip()
        if not line:
            continue
        if _NOISE_LINE_PREFIX_RE.search(line):
            continue
        line = re.sub(r"^\s*FBA\s*[:：]\s*", "", line, flags=re.IGNORECASE).strip()
        if not line:
            continue
        # Do not drop the whole line for one-line OCR; cut known footer tails instead.
        line = re.split(r"\b(?:CREATED|ERSTELLT)\b", line, maxsplit=1, flags=re.IGNORECASE)[0].strip()
        line = re.split(r"\bFBA[A-Z0-9]{6,}\b", line, maxsplit=1, flags=re.IGNORECASE)[0].strip()
        if not line:
            continue
        if _is_country_line(line):
            continue
        lines.append(line)
    return lines


def _has_required_fields(record: _AddressRecord, country_mode: str) -> bool:
    mode = (country_mode or "").upper()
    if mode == "US":
        return bool(record.street_no and record.city and record.state and record.zip5)
    if mode == "DE":
        return bool(record.street_no and record.street and record.city and record.zip5)
    if mode == "UK":
        return bool(record.postal_code and record.city and record.street)
    if mode == "AU":
        return bool(record.postal_code and record.state and record.city and record.street)
    return False


def _missing_required_fields(record: _AddressRecord, country_mode: str) -> List[str]:
    mode = (country_mode or "").upper()
    fields: List[str] = []
    if mode == "US":
        checks = ("street_no", "city", "state", "zip5")
    elif mode == "DE":
        checks = ("street_no", "street", "city", "zip5")
    elif mode == "UK":
        checks = ("postal_code", "city", "street")
    elif mode == "AU":
        checks = ("postal_code", "state", "city", "street")
    else:
        checks = ()
    for field_name in checks:
        if not str(getattr(record, field_name, "")).strip():
            fields.append(field_name)
    return fields


def _extract_street_segment(text: str, country_mode: str) -> str:
    source = _normalize_street_text(text, country_mode)
    if not source:
        return source

    if country_mode in ("US", "UK", "AU"):
        road_keywords_pattern = "|".join(_ROAD_KEYWORDS)
        match = re.search(
            rf"(\d{{1,5}}\s*[A-Z0-9\s\-]{{0,90}}\b(?:{road_keywords_pattern})\b)",
            source,
        )
        if match:
            return _normalize_spaces(match.group(1))

    if country_mode == "DE":
        match = re.search(r"([A-Z][A-Z\s\-]*STRASSE)\s*(\d+[A-Z]?)\b", source)
        if match:
            return _normalize_spaces(f"{match.group(1)} {match.group(2)}")
        match = re.search(r"([A-Z0-9\s\-]{1,90}\s+\d+[A-Z]?)$", source)
        if match:
            return _normalize_spaces(match.group(1))

    return source


def _extract_us_city_state_zip(text: str) -> Optional[Tuple[str, str, str]]:
    raw = str(text or "").upper()
    raw = (
        raw.replace("Ä", "AE")
        .replace("Ö", "OE")
        .replace("Ü", "UE")
        .replace("ß", "SS")
    )
    raw = re.sub(r"[^A-Z0-9,\s\-]", " ", raw)
    raw = _normalize_spaces(raw)
    if not raw:
        return None

    road_stop_tokens = set(_ROAD_KEYWORDS) | {"APT", "SUITE", "STE", "UNIT", "BLDG", "BUILDING", "NO"}
    state_zip_matches = list(_US_STATE_ZIP_RE.finditer(raw))
    candidates: List[Tuple[str, str, str, int]] = []
    for match in state_zip_matches:
        state = match.group(1).strip()
        zip5 = match.group(2)[:5]
        prefix = raw[:match.start()].strip()
        if not prefix:
            continue

        tail = prefix.split(",")[-1].strip() if "," in prefix else prefix
        tokens = tail.split()
        if not tokens:
            continue
        city_tokens_reversed: List[str] = []
        for token in reversed(tokens):
            if re.search(r"\d", token):
                break
            if token in _COUNTRY_WORD_TOKENS:
                if city_tokens_reversed:
                    break
                continue
            if token in road_stop_tokens:
                if city_tokens_reversed:
                    break
                continue
            city_tokens_reversed.append(token)
            if len(city_tokens_reversed) >= 3:
                break
        if not city_tokens_reversed:
            continue
        city = " ".join(reversed(city_tokens_reversed))
        candidates.append((city, state, zip5, match.start()))

    if candidates:
        candidates.sort(key=lambda item: item[3], reverse=True)
        return candidates[0][0], candidates[0][1], candidates[0][2]

    line = _normalize_text_for_match(raw)
    regex_matches = list(_US_CITY_STATE_ZIP_RE.finditer(line))
    if not regex_matches:
        return None

    filtered: List[Tuple[str, str, str, int, int]] = []
    for match in regex_matches:
        city = _normalize_spaces(match.group(1))
        state = match.group(2).strip()
        zip5 = match.group(3)[:5]
        city_tokens = city.split()
        if not city_tokens:
            continue
        if any(token in _ROAD_KEYWORDS for token in city_tokens):
            continue
        if any(re.search(r"\d", token) for token in city_tokens):
            continue
        filtered.append((city, state, zip5, len(city_tokens), match.start()))

    if not filtered:
        last = regex_matches[-1]
        return _normalize_spaces(last.group(1)), last.group(2).strip(), last.group(3)[:5]

    latest_start = max(item[4] for item in filtered)
    latest_candidates = [item for item in filtered if item[4] == latest_start]
    latest_candidates.sort(key=lambda item: item[3])
    chosen = latest_candidates[0]
    return chosen[0], chosen[1], chosen[2]


def _extract_uk_city_postal(text: str) -> Optional[Tuple[str, str]]:
    line = _normalize_text_for_match(text)
    matches = list(_UK_POSTAL_RE.finditer(line))
    if not matches:
        return None
    match = matches[-1]
    postal = _normalize_uk_postal_code(match.group(1))
    if not postal:
        return None

    prefix = _normalize_spaces(line[:match.start()])
    tokens = prefix.split()
    city_tokens_reversed: List[str] = []
    for token in reversed(tokens):
        if re.search(r"\d", token):
            break
        if token in _COUNTRY_WORD_TOKENS:
            if city_tokens_reversed:
                break
            continue
        if token in _ROAD_KEYWORDS:
            if city_tokens_reversed:
                break
            continue
        if token in _SOURCE_NOISE_TOKENS:
            if city_tokens_reversed:
                break
            continue
        if token in _UK_REGION_TOKENS:
            if city_tokens_reversed:
                break
            continue
        city_tokens_reversed.append(token)
        if len(city_tokens_reversed) >= 2:
            break
    if not city_tokens_reversed:
        return "", postal
    city = " ".join(reversed(city_tokens_reversed))
    return city, postal


def _extract_uk_street(lines: List[str], city: str, postal: str) -> str:
    road_pattern = re.compile(r"\b([A-Z0-9][A-Z0-9\s\-]{0,48}?\b(?:ROAD|STREET|DRIVE|LANE|WAY|AVENUE|PARK))\b")
    city_pattern = re.compile(rf"\b{re.escape(_normalize_text_for_match(city))}\b") if city else None
    postal_spaced = _normalize_text_for_match(postal)
    postal_compact = re.sub(r"[^A-Z0-9]", "", postal_spaced)
    segments: List[str] = []

    for raw in lines:
        text = _normalize_street_text(raw, "UK")
        if not text:
            continue
        if city_pattern:
            split_parts = city_pattern.split(text, maxsplit=1)
            text = split_parts[0] if split_parts else text
        if postal_spaced:
            text = re.sub(rf"\b{re.escape(postal_spaced)}\b", " ", text)
        if postal_compact:
            text = re.sub(rf"\b{re.escape(postal_compact)}\b", " ", re.sub(r"[^A-Z0-9\s]", " ", text))
            text = _normalize_street_text(text, "UK")
        for match in road_pattern.finditer(text):
            segment = _normalize_spaces(match.group(1))
            if not segment:
                continue
            raw_tokens = segment.split()
            cleaned_tokens: List[str] = []
            for token in raw_tokens:
                if token in _SOURCE_NOISE_TOKENS:
                    continue
                if token in _UK_STREET_DROP_TOKENS:
                    continue
                if re.fullmatch(r"\d{5,}", token):
                    continue
                cleaned_tokens.append(token)
            cleaned_segment = _normalize_spaces(" ".join(cleaned_tokens))
            if not cleaned_segment:
                continue
            if not re.search(r"\b(ROAD|STREET|DRIVE|LANE|WAY|AVENUE|PARK)\b", cleaned_segment):
                continue
            segments.append(cleaned_segment)

    if not segments:
        return ""
    return " ".join(_unique_preserve_order(segments))


def _format_address_record_for_reply(record: _AddressRecord) -> str:
    parts: List[str] = []
    if record.company:
        parts.append(f"公司={record.company}")
    if record.street:
        parts.append(f"街道={record.street}")
    if record.street_no:
        parts.append(f"门牌号={record.street_no}")
    if record.city:
        parts.append(f"城市={record.city}")
    if record.state:
        parts.append(f"州={record.state}")
    postal = record.postal_code or record.zip5
    if postal:
        parts.append(f"邮编={postal}")
    if record.fc_code:
        parts.append(f"仓库编码={record.fc_code}")
    return "; ".join(parts) if parts else "-"


def _format_ocr_record_for_reply(record: _AddressRecord) -> str:
    lines: List[str] = []
    if record.fc_code:
        lines.append(f"仓库编码={record.fc_code};")
    if record.street:
        lines.append(f"街道={record.street};")
    if record.street_no:
        lines.append(f"门牌号={record.street_no};")
    if record.city:
        lines.append(f"城市={record.city};")
    if record.state:
        lines.append(f"州={record.state};")
    postal = record.postal_code or record.zip5
    if postal:
        lines.append(f"邮编={postal};")
    return "\n".join(lines) if lines else "-"


def _parse_address_record(text: str, country_mode: str, fc_code: str = "") -> _AddressRecord:
    lines = _prepare_address_lines(text)
    normalized_lines = [_normalize_text_for_match(line) for line in lines]
    if country_mode == "DE":
        normalized_lines = [
            _normalize_spaces(re.sub(r"(?<=\d)(?=[A-Z])", " ", re.sub(r"(?<=[A-Z])(?=\d)", " ", line)))
            for line in normalized_lines
        ]
    record = _AddressRecord(
        country_mode=country_mode,
        raw_text=str(text or ""),
        fc_code=_normalize_fc_code(fc_code),
    )
    if not normalized_lines:
        return record

    used_indexes = set()
    city_line_index: Optional[int] = None

    street_line_index: Optional[int] = None
    if country_mode == "DE":
        for idx, line in enumerate(normalized_lines):
            if "STRASSE" in line and re.search(r"\d+[A-Z]?$", line):
                street_line_index = idx
                break
        if street_line_index is None:
            for idx, line in enumerate(normalized_lines):
                if re.search(r"[A-Z].*\d+[A-Z]?$", line):
                    street_line_index = idx
                    break
    else:
        for idx, line in enumerate(normalized_lines):
            if country_mode in ("US", "UK", "AU") and _is_source_noise_line(line):
                if country_mode == "US" and _US_STATE_ZIP_RE.search(line):
                    pass
                elif country_mode == "UK" and _UK_POSTAL_RE.search(line):
                    pass
                elif country_mode == "AU" and _AU_CITY_STATE_POSTAL_RE.search(line):
                    pass
                else:
                    continue
            contains_road_keyword = any(keyword in line for keyword in _ROAD_KEYWORDS)
            has_number = bool(re.search(r"\d", line))
            if country_mode in ("UK", "AU") and contains_road_keyword:
                street_line_index = idx
                break
            if country_mode == "US" and has_number and contains_road_keyword:
                street_line_index = idx
                break
    if street_line_index is None:
        for idx, line in enumerate(normalized_lines):
            if country_mode in ("US", "UK", "AU") and _is_source_noise_line(line):
                if country_mode == "US" and _US_STATE_ZIP_RE.search(line):
                    pass
                elif country_mode == "UK" and _UK_POSTAL_RE.search(line):
                    pass
                elif country_mode == "AU" and _AU_CITY_STATE_POSTAL_RE.search(line):
                    pass
                else:
                    continue
            if re.search(r"\d", line):
                street_line_index = idx
                break
    if street_line_index is not None:
        used_indexes.add(street_line_index)
        street_segment = _extract_street_segment(normalized_lines[street_line_index], country_mode)
        record.street = _normalize_street_text(street_segment, country_mode)
        record.street_no = _extract_street_no(record.street, country_mode)

    if country_mode == "US":
        for idx, line in enumerate(normalized_lines):
            parsed = _extract_us_city_state_zip(line)
            if not parsed:
                continue
            city_line_index = idx
            used_indexes.add(idx)
            record.city, record.state, record.zip5 = parsed
            record.postal_code = record.zip5
            break

    elif country_mode == "DE":
        for idx, line in enumerate(normalized_lines):
            match = _DE_ZIP_CITY_RE.search(line)
            if not match:
                continue
            city_line_index = idx
            used_indexes.add(idx)
            record.zip5 = match.group(1)
            record.postal_code = record.zip5
            city_raw = _normalize_spaces(match.group(2))
            city_tokens: List[str] = []
            for token in city_raw.split():
                if token in _DE_CITY_STOP_TOKENS:
                    break
                if re.search(r"\d", token):
                    break
                city_tokens.append(token)
                if len(city_tokens) >= 3:
                    break
            record.city = " ".join(city_tokens) if city_tokens else city_raw.split()[0]
            break

    elif country_mode == "UK":
        for idx, line in enumerate(normalized_lines):
            parsed = _extract_uk_city_postal(line)
            if not parsed:
                continue
            city_line_index = idx
            used_indexes.add(idx)
            parsed_city, parsed_postal = parsed
            record.postal_code = parsed_postal
            if parsed_city:
                record.city = parsed_city
            break
        if not record.city:
            start_idx = (city_line_index - 1) if city_line_index is not None else (len(normalized_lines) - 1)
            for idx in range(start_idx, -1, -1):
                candidate = normalized_lines[idx]
                if re.search(r"\d", candidate) or _is_country_line(candidate):
                    continue
                tokens = [t for t in candidate.split() if t not in _UK_REGION_TOKENS and t not in _COUNTRY_WORD_TOKENS]
                if not tokens:
                    continue
                record.city = " ".join(tokens[-2:]) if len(tokens) >= 2 else tokens[0]
                used_indexes.add(idx)
                break
        uk_street = _extract_uk_street(normalized_lines, record.city, record.postal_code)
        if uk_street:
            record.street = uk_street
            record.street_no = _extract_street_no(record.street, country_mode)

    elif country_mode == "AU":
        for idx, line in enumerate(normalized_lines):
            match = _AU_CITY_STATE_POSTAL_RE.search(line)
            if not match:
                continue
            city_line_index = idx
            used_indexes.add(idx)
            record.city = _normalize_spaces(match.group(1))
            record.state = _normalize_au_state(match.group(2))
            record.postal_code = match.group(3)
            record.zip5 = record.postal_code
            break

    company_lines: List[str] = []
    for idx, line in enumerate(normalized_lines):
        if idx in used_indexes:
            continue
        if re.search(r"\d", line):
            continue
        if _is_country_line(line):
            continue
        company_lines.append(line)
    record.company = _normalize_spaces(" ".join(company_lines))
    return record


def _score_address_candidate(country_mode: str, source: _AddressRecord, candidate: _AddressRecord) -> _AddressCandidateScore:
    mode = (country_mode or "").upper()
    street_sim = _street_similarity(source.street, candidate.street)
    company_sim = _company_similarity(source.company, candidate.company)
    hard_ok = True
    hard_reasons: List[str] = []
    score = 0.0

    if mode == "US":
        zip_ok = source.zip5 == candidate.zip5 and bool(source.zip5)
        city_ok = _city_matches(source.city, candidate.city)
        state_ok = source.state == candidate.state and bool(source.state)
        street_no_ok = source.street_no == candidate.street_no and bool(source.street_no)
        hard_ok = zip_ok and city_ok and state_ok and street_no_ok
        if not zip_ok:
            hard_reasons.append("zip5")
        if not city_ok:
            hard_reasons.append("city")
        if not state_ok:
            hard_reasons.append("state")
        if not street_no_ok:
            hard_reasons.append("street_no")
        score += 25.0 if zip_ok else 0.0
        score += 20.0 if city_ok else 0.0
        score += 20.0 if state_ok else 0.0
        score += 20.0 if street_no_ok else 0.0
        score += 10.0 * street_sim
        score += 5.0 * company_sim

    elif mode == "DE":
        zip_ok = source.zip5 == candidate.zip5 and bool(source.zip5)
        city_ok = _city_matches(source.city, candidate.city)
        street_no_ok = source.street_no == candidate.street_no and bool(source.street_no)
        street_ok = street_sim >= 0.80
        hard_ok = zip_ok and city_ok and street_no_ok and street_ok
        if not zip_ok:
            hard_reasons.append("zip5")
        if not city_ok:
            hard_reasons.append("city")
        if not street_no_ok:
            hard_reasons.append("street_no")
        if not street_ok:
            hard_reasons.append("street")
        score += 35.0 if zip_ok else 0.0
        score += 25.0 if city_ok else 0.0
        score += 20.0 if street_no_ok else 0.0
        score += 15.0 * street_sim
        score += 5.0 * company_sim

    elif mode == "UK":
        postal_ok = source.postal_code == candidate.postal_code and bool(source.postal_code)
        city_ok = source.city == candidate.city and bool(source.city)
        uk_street_coverage = _uk_street_token_coverage(source.street, candidate.street)
        street_ok = street_sim >= 0.60 or uk_street_coverage >= 0.80
        hard_ok = postal_ok and city_ok and street_ok
        if not postal_ok:
            hard_reasons.append("postal_code")
        if not city_ok:
            hard_reasons.append("city")
        if not street_ok:
            hard_reasons.append("street")
        score += 40.0 if postal_ok else 0.0
        score += 25.0 if city_ok else 0.0
        score += 25.0 * max(street_sim, uk_street_coverage)
        score += 10.0 * company_sim

    elif mode == "AU":
        postal_ok = source.postal_code == candidate.postal_code and bool(source.postal_code)
        state_ok = source.state == candidate.state and bool(source.state)
        city_ok = source.city == candidate.city and bool(source.city)
        street_ok = street_sim >= 0.75
        hard_ok = postal_ok and state_ok and city_ok and street_ok
        if not postal_ok:
            hard_reasons.append("postal_code")
        if not state_ok:
            hard_reasons.append("state")
        if not city_ok:
            hard_reasons.append("city")
        if not street_ok:
            hard_reasons.append("street")
        score += 35.0 if postal_ok else 0.0
        score += 20.0 if state_ok else 0.0
        score += 20.0 if city_ok else 0.0
        score += 20.0 * street_sim
        score += 5.0 * company_sim

    else:
        hard_ok = False
        hard_reasons.append("unsupported_country")

    return _AddressCandidateScore(
        record=candidate,
        score=round(score, 2),
        hard_ok=hard_ok,
        hard_reason=",".join(hard_reasons) if hard_reasons else "",
    )


def _extract_download_payload(response: Dict[str, Any]) -> Tuple[Optional[str], Optional[bytes], Optional[str]]:
    data = response.get("data") or response.get("Data") or response
    if isinstance(data, str):
        if data.startswith("http"):
            return data, None, None
        return None, None, None
    if isinstance(data, dict):
        url = (
            data.get("file_url")
            or data.get("fileUrl")
            or data.get("download_url")
            or data.get("downloadUrl")
            or data.get("url")
        )
        file_name = data.get("file_name") or data.get("fileName")
        content_b64 = (
            data.get("file_content")
            or data.get("fileContent")
            or data.get("content")
            or data.get("base64")
            or data.get("file_base64")
        )
        if content_b64:
            try:
                raw = base64.b64decode(content_b64)
                return url, raw, file_name
            except Exception:
                pass
        return url, None, file_name
    return None, None, None


def _is_param_error(response: Dict[str, Any]) -> bool:
    code = response.get("code") or response.get("Code") or response.get("status")
    msg = response.get("message") or response.get("msg") or response.get("Message") or ""
    return str(code) == "102" or "参数错误" in str(msg)


def _extract_message_id(incoming_message, raw_data) -> Optional[str]:
    for attr in ("message_id", "messageId", "msg_id", "msgId", "msgid"):
        if hasattr(incoming_message, attr):
            value = getattr(incoming_message, attr)
            if value:
                return str(value)
    if isinstance(raw_data, str):
        try:
            raw_data = json.loads(raw_data)
        except Exception:
            raw_data = None
    if isinstance(raw_data, dict):
        for key in ("messageId", "message_id", "msgId", "msg_id"):
            value = raw_data.get(key)
            if value:
                return str(value)
        nested = raw_data.get("data")
        if isinstance(nested, dict):
            for key in ("messageId", "message_id", "msgId", "msg_id"):
                value = nested.get(key)
                if value:
                    return str(value)
    return None


class ShipmentBusyError(RuntimeError):
    """Raised when the same shipment is already being processed."""


@dataclass
class _QueuedJob:
    request_id: str
    message_id: Optional[str]
    raw_data: Any
    sender_user_id: Optional[str]
    shipment_sns: List[str]
    queued_at: float
    selected_file: Optional["_SelectedFileChoice"] = None


@dataclass
class _ReplyResult:
    summary_text: Optional[str]
    shipment_messages: List[str]
    shipment_images: List[Path]
    selection_messages: List[str]
    pending_file_selections: List["_PendingFileSelection"]


@dataclass
class _SelectableFileOption:
    file_id: str
    file_name: str


@dataclass
class _PendingFileSelection:
    shipment_sn: str
    nation: Optional[str]
    dest_ids: List[str]
    options: List[_SelectableFileOption]
    expires_at: float


@dataclass
class _SelectedFileChoice:
    shipment_sn: str
    nation: Optional[str]
    dest_ids: List[str]
    file_id: str
    file_name: str


class _StateStore:
    MESSAGE_DEDUP_TABLE = "dim_bot_cp_message_dedup"
    SHIPMENT_LOCKS_TABLE = "dim_bot_cp_shipment_lock"
    REQUEST_LOG_TABLE = "fact_bot_cp_call_log"
    REQUIRED_TABLES = (MESSAGE_DEDUP_TABLE, SHIPMENT_LOCKS_TABLE, REQUEST_LOG_TABLE)

    def __init__(
        self,
        *,
        host: str,
        port: int,
        user: str,
        password: str,
        database: str,
        connect_timeout_sec: int = 5,
    ):
        self.host = host
        self.port = port
        self.user = user
        self.password = password
        self.database = database
        self.connect_timeout_sec = connect_timeout_sec

        self._lock = threading.Lock()
        self._conn: Optional[pymysql.connections.Connection] = None
        self._connect()
        self._assert_schema_ready()

    def _connect(self) -> None:
        self._conn = pymysql.connect(
            host=self.host,
            port=self.port,
            user=self.user,
            password=self.password,
            database=self.database,
            charset="utf8mb4",
            autocommit=False,
            connect_timeout=self.connect_timeout_sec,
            read_timeout=30,
            write_timeout=30,
        )

    def _ensure_conn(self) -> pymysql.connections.Connection:
        if self._conn is None:
            self._connect()
            assert self._conn is not None
            return self._conn
        try:
            self._conn.ping(reconnect=True)
        except Exception:
            self._connect()
        assert self._conn is not None
        return self._conn

    def _assert_schema_ready(self) -> None:
        conn = self._ensure_conn()
        placeholders = ", ".join(["%s"] * len(self.REQUIRED_TABLES))
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = %s
                  AND table_name IN ({placeholders})
                """,
                (self.database, *self.REQUIRED_TABLES),
            )
            rows = cur.fetchall()
        existing = {str(row[0]) for row in rows}
        missing = [table for table in self.REQUIRED_TABLES if table not in existing]
        if missing:
            raise RuntimeError(
                "MySQL tables missing: "
                + ", ".join(missing)
                + ". Please run sql/mysql_state.sql manually before starting the bot."
            )

    @staticmethod
    def _normalize_sns(shipment_sns: List[str]) -> List[str]:
        return sorted({str(sn).strip().upper() for sn in shipment_sns if str(sn).strip()})

    @classmethod
    def _cleanup_expired_tx(cls, cur: pymysql.cursors.Cursor, now_ts: float) -> None:
        cur.execute(f"DELETE FROM {cls.MESSAGE_DEDUP_TABLE} WHERE expires_at <= %s", (now_ts,))
        cur.execute(f"DELETE FROM {cls.SHIPMENT_LOCKS_TABLE} WHERE expires_at <= %s", (now_ts,))

    def register_message(self, message_id: str, ttl_sec: int) -> bool:
        now_ts = time.time()
        expires_at = now_ts + ttl_sec
        with self._lock:
            conn = self._ensure_conn()
            cur = conn.cursor()
            try:
                conn.begin()
                self._cleanup_expired_tx(cur, now_ts)
                cur.execute(
                    f"SELECT expires_at FROM {self.MESSAGE_DEDUP_TABLE} WHERE message_id = %s FOR UPDATE",
                    (message_id,),
                )
                row = cur.fetchone()
                if row and float(row[0]) > now_ts:
                    conn.rollback()
                    return False
                cur.execute(
                    f"""
                    INSERT INTO {self.MESSAGE_DEDUP_TABLE}(message_id, expires_at)
                    VALUES(%s, %s)
                    ON DUPLICATE KEY UPDATE expires_at = VALUES(expires_at)
                    """,
                    (message_id, expires_at),
                )
                conn.commit()
                return True
            except Exception:
                conn.rollback()
                raise
            finally:
                cur.close()

    def acquire_shipment_locks(self, shipment_sns: List[str], holder_id: str, ttl_sec: int) -> Tuple[bool, Optional[str]]:
        keys = self._normalize_sns(shipment_sns)
        if not keys:
            return True, None
        now_ts = time.time()
        expires_at = now_ts + ttl_sec
        with self._lock:
            conn = self._ensure_conn()
            cur = conn.cursor()
            try:
                conn.begin()
                self._cleanup_expired_tx(cur, now_ts)
                for key in keys:
                    cur.execute(
                        f"SELECT holder_id, expires_at FROM {self.SHIPMENT_LOCKS_TABLE} WHERE shipment_sn = %s FOR UPDATE",
                        (key,),
                    )
                    row = cur.fetchone()
                    if row and row[0] != holder_id and float(row[1]) > now_ts:
                        conn.rollback()
                        return False, key
                for key in keys:
                    cur.execute(
                        f"""
                        INSERT INTO {self.SHIPMENT_LOCKS_TABLE}(shipment_sn, holder_id, expires_at)
                        VALUES(%s, %s, %s)
                        ON DUPLICATE KEY UPDATE holder_id = VALUES(holder_id), expires_at = VALUES(expires_at)
                        """,
                        (key, holder_id, expires_at),
                    )
                conn.commit()
                return True, None
            except Exception:
                conn.rollback()
                raise
            finally:
                cur.close()

    def release_shipment_locks(self, shipment_sns: List[str], holder_id: str) -> None:
        keys = self._normalize_sns(shipment_sns)
        if not keys:
            return
        placeholders = ",".join("%s" for _ in keys)
        params = [holder_id, *keys]
        with self._lock:
            conn = self._ensure_conn()
            with conn.cursor() as cur:
                try:
                    conn.begin()
                    cur.execute(
                        f"DELETE FROM {self.SHIPMENT_LOCKS_TABLE} WHERE holder_id = %s AND shipment_sn IN ({placeholders})",
                        params,
                    )
                    conn.commit()
                except Exception:
                    conn.rollback()
                    raise

    def log_request_event(
        self,
        *,
        request_id: str,
        message_id: Optional[str],
        user_id: Optional[str],
        user_name: Optional[str],
        event_type: str,
        ack_status: Optional[str],
        shipment_sns: Optional[List[str]],
        detail: Optional[str],
    ) -> None:
        if str(event_type).strip().upper() != "RECEIVED":
            return
        safe_user_id = (str(user_id or "").strip() or "UNKNOWN")[:64]
        safe_user_name = (str(user_name or "").strip() or None)
        if safe_user_name:
            safe_user_name = safe_user_name[:128]
        message_text = (detail or "").strip()
        if shipment_sns:
            sn_text = ",".join(self._normalize_sns(shipment_sns))
            if sn_text:
                message_text = f"{message_text}\n\nshipment_sns={sn_text}" if message_text else f"shipment_sns={sn_text}"
        with self._lock:
            conn = self._ensure_conn()
            with conn.cursor() as cur:
                try:
                    conn.begin()
                    cur.execute(
                        f"""
                        INSERT INTO {self.REQUEST_LOG_TABLE}
                        (user_id, user_name, message_text)
                        VALUES(%s, %s, %s)
                        """,
                        (
                            safe_user_id,
                            safe_user_name,
                            message_text or None,
                        ),
                    )
                    conn.commit()
                except Exception:
                    conn.rollback()
                    raise


class ShipmentQueryHandler(dingtalk_stream.ChatbotHandler):
    def __init__(self, logger: Optional[logging.Logger] = None):
        super().__init__()
        self.logger = logger or logging.getLogger(__name__)
        self.client = LingXingClient(
            host=config.LINGXING_API_HOST,
            app_id=config.LINGXING_API_KEY,
            app_secret=config.LINGXING_API_SECRET,
            token_url=config.LINGXING_TOKEN_URL,
            token_key=config.LINGXING_TOKEN_REQUEST_KEY,
            ssl_verify=config.LINGXING_SSL_VERIFY,
        )
        self._ocr_client: Optional[AliyunOCRClient] = None
        self._ocr_client_lock = threading.Lock()

        self._max_concurrent_requests = max(1, int(config.MAX_CONCURRENT_REQUESTS))
        self._message_dedup_ttl_sec = max(60, int(config.MESSAGE_DEDUP_TTL_SEC))
        self._shipment_lock_ttl_sec = max(60, int(config.SHIPMENT_LOCK_TTL_SEC))
        self._download_concurrency = max(1, int(config.DOWNLOAD_CONCURRENCY))
        self._ocr_concurrency = max(1, int(config.OCR_CONCURRENCY))
        self._resource_wait_timeout_sec = max(5, int(config.RESOURCE_WAIT_TIMEOUT_SEC))
        self._queue_max_size = max(self._max_concurrent_requests, int(config.JOB_QUEUE_MAX_SIZE))
        self._retry_times = max(1, int(config.API_RETRY_TIMES))
        self._retry_base_delay_sec = max(0.1, float(config.API_RETRY_BASE_DELAY_SEC))
        self._retry_max_delay_sec = max(self._retry_base_delay_sec, float(config.API_RETRY_MAX_DELAY_SEC))
        self._download_retention_days = max(1, int(config.DOWNLOAD_RETENTION_DAYS))
        self._download_cleanup_interval_sec = max(60, int(config.DOWNLOAD_CLEANUP_INTERVAL_SEC))

        missing_db = [name for name, value in {
            "DB_HOST": config.DB_HOST,
            "DB_USER": config.DB_USER,
            "DB_NAME": config.DB_NAME,
        }.items() if not str(value).strip()]
        if missing_db:
            raise RuntimeError(f"Missing DB config: {', '.join(missing_db)}")

        self._state_store = _StateStore(
            host=config.DB_HOST,
            port=config.DB_PORT,
            user=config.DB_USER,
            password=config.DB_PASSWORD,
            database=config.DB_NAME,
            connect_timeout_sec=config.DB_CONNECT_TIMEOUT_SEC,
        )
        self._job_queue: asyncio.Queue[_QueuedJob] = asyncio.Queue(maxsize=self._queue_max_size)
        self._worker_tasks: List[asyncio.Task] = []
        self._workers_started = False
        self._worker_init_lock = asyncio.Lock()
        self._cleanup_lock = threading.Lock()
        self._last_cleanup_monotonic = 0.0

        self._download_limiter = threading.BoundedSemaphore(self._download_concurrency)
        self._ocr_limiter = threading.BoundedSemaphore(self._ocr_concurrency)
        self._address_book_lock = threading.Lock()
        self._address_book_loaded = False
        self._address_book_disabled_reason: Optional[str] = None
        self._address_book_by_country_port: Dict[str, Dict[str, List[_AddressRecord]]] = {}
        self._pending_file_selection_lock = threading.Lock()
        self._pending_file_selection_ttl_sec = 1800
        self._pending_file_selections: Dict[Tuple[str, str], _PendingFileSelection] = {}
        self._tech_user_ids: List[str] = list(config.DING_TECH_USER_IDS)
        if self._tech_user_ids:
            self.logger.info("Tech alert enabled for %s user(s).", len(self._tech_user_ids))

    def _build_tech_alert_text(
        self,
        *,
        request_id: Optional[str],
        stage: str,
        exc: Exception,
        shipment_sns: Optional[List[str]] = None,
    ) -> str:
        trace_text = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__)).strip()
        if len(trace_text) > 1800:
            trace_text = trace_text[-1800:]
        lines = [
            "[CP机器人异常告警]",
            f"request_id: {request_id or '-'}",
            f"stage: {stage}",
            f"error: {type(exc).__name__}: {exc}",
        ]
        if shipment_sns:
            lines.append(f"shipment_sns: {','.join(shipment_sns)}")
        if trace_text:
            lines.append("")
            lines.append(trace_text)
        return "\n".join(lines)

    def _notify_tech_exception_sync(
        self,
        *,
        request_id: Optional[str],
        stage: str,
        exc: Exception,
        shipment_sns: Optional[List[str]] = None,
    ) -> None:
        if not self._tech_user_ids:
            return
        alert_text = self._build_tech_alert_text(
            request_id=request_id,
            stage=stage,
            exc=exc,
            shipment_sns=shipment_sns,
        )
        for tech_user_id in self._tech_user_ids:
            try:
                _DEFAULT_NOTIFIER.send_user_text(tech_user_id, alert_text)
            except Exception as notify_exc:
                self.logger.warning(
                    "[req=%s] Tech alert send failed(user=%s): %s",
                    request_id or "-",
                    tech_user_id,
                    notify_exc,
                )

    async def _notify_tech_exception(
        self,
        *,
        request_id: Optional[str],
        stage: str,
        exc: Exception,
        shipment_sns: Optional[List[str]] = None,
    ) -> None:
        if not self._tech_user_ids:
            return
        alert_text = self._build_tech_alert_text(
            request_id=request_id,
            stage=stage,
            exc=exc,
            shipment_sns=shipment_sns,
        )
        for tech_user_id in self._tech_user_ids:
            try:
                await asyncio.to_thread(_DEFAULT_NOTIFIER.send_user_text, tech_user_id, alert_text)
            except Exception as notify_exc:
                self.logger.warning(
                    "[req=%s] Tech alert send failed(user=%s): %s",
                    request_id or "-",
                    tech_user_id,
                    notify_exc,
                )

    async def _ensure_workers(self) -> None:
        if self._workers_started:
            return
        async with self._worker_init_lock:
            if self._workers_started:
                return
            for index in range(self._max_concurrent_requests):
                worker_id = index + 1
                task = asyncio.create_task(self._worker_loop(worker_id), name=f"cp-bot-worker-{worker_id}")
                self._worker_tasks.append(task)
            self._workers_started = True
            self.logger.info("Started %s worker(s) for async queue processing.", self._max_concurrent_requests)

    @contextmanager
    def _acquire_limiter(self, limiter: threading.BoundedSemaphore, name: str):
        acquired = limiter.acquire(timeout=self._resource_wait_timeout_sec)
        if not acquired:
            raise RuntimeError(f"{name} busy: wait timeout {self._resource_wait_timeout_sec}s")
        try:
            yield
        finally:
            limiter.release()

    def _register_message_once(self, message_id: str, request_id: Optional[str] = None) -> bool:
        try:
            return self._state_store.register_message(message_id, self._message_dedup_ttl_sec)
        except Exception as exc:
            self.logger.warning("State store dedup failed, fallback allow this message: %s", exc)
            self._notify_tech_exception_sync(
                request_id=request_id,
                stage="state_store_register_message",
                exc=exc,
            )
            return True

    def _acquire_shipment_locks(self, shipment_sns: List[str], request_id: str) -> Tuple[bool, Optional[str]]:
        try:
            return self._state_store.acquire_shipment_locks(shipment_sns, request_id, self._shipment_lock_ttl_sec)
        except Exception as exc:
            self.logger.warning("State store lock acquire failed, fallback allow this request: %s", exc)
            self._notify_tech_exception_sync(
                request_id=request_id,
                stage="state_store_acquire_lock",
                exc=exc,
                shipment_sns=shipment_sns,
            )
            return True, None

    def _release_shipment_locks(self, shipment_sns: List[str], request_id: str) -> None:
        try:
            self._state_store.release_shipment_locks(shipment_sns, request_id)
        except Exception as exc:
            self.logger.warning("[req=%s] State store lock release failed: %s", request_id, exc)
            self._notify_tech_exception_sync(
                request_id=request_id,
                stage="state_store_release_lock",
                exc=exc,
                shipment_sns=shipment_sns,
            )

    def _cleanup_pending_file_selections(self) -> None:
        now_ts = time.time()
        expired_keys: List[Tuple[str, str]] = []
        for key, pending in self._pending_file_selections.items():
            if pending.expires_at <= now_ts:
                expired_keys.append(key)
        for key in expired_keys:
            self._pending_file_selections.pop(key, None)

    def _register_pending_file_selection(self, user_id: str, pending: _PendingFileSelection) -> None:
        key = (str(user_id).strip(), str(pending.shipment_sn).strip().upper())
        if not key[0] or not key[1]:
            return
        with self._pending_file_selection_lock:
            self._cleanup_pending_file_selections()
            self._pending_file_selections[key] = pending

    def _consume_pending_file_selection(
        self,
        user_id: str,
        shipment_sn: str,
        index: int,
    ) -> Tuple[Optional[_SelectedFileChoice], Optional[str]]:
        user_key = str(user_id).strip()
        sn_key = str(shipment_sn).strip().upper()
        with self._pending_file_selection_lock:
            self._cleanup_pending_file_selections()
            pending = self._pending_file_selections.get((user_key, sn_key))
            if not pending:
                return None, "未找到待选择文件，请重新发送发货单号。"
            if index > len(pending.options):
                return None, f"序号超出范围，可选范围为 1~{len(pending.options)}。"
            selected = pending.options[index - 1]
            self._pending_file_selections.pop((user_key, sn_key), None)
        return _SelectedFileChoice(
            shipment_sn=sn_key,
            nation=pending.nation,
            dest_ids=list(pending.dest_ids),
            file_id=selected.file_id,
            file_name=selected.file_name,
        ), None

    def _consume_pending_file_selection_by_name(
        self,
        user_id: str,
        file_name_text: str,
    ) -> Tuple[Optional[_SelectedFileChoice], Optional[str]]:
        user_key = str(user_id).strip()
        target_name = Path(str(file_name_text or "").strip()).name.strip().lower()
        if not user_key or not target_name:
            return None, None

        matches: List[Tuple[Tuple[str, str], _PendingFileSelection, _SelectableFileOption]] = []
        with self._pending_file_selection_lock:
            self._cleanup_pending_file_selections()
            for key, pending in self._pending_file_selections.items():
                if key[0] != user_key:
                    continue
                for option in pending.options:
                    option_name = Path(str(option.file_name or "")).name.strip().lower()
                    if option_name and option_name == target_name:
                        matches.append((key, pending, option))

            if not matches:
                return None, None
            if len(matches) > 1:
                return None, "匹配到多个同名文件，请用“选择文件 SP单号 序号”进行选择。"

            key, pending, selected = matches[0]
            self._pending_file_selections.pop(key, None)

        return _SelectedFileChoice(
            shipment_sn=pending.shipment_sn,
            nation=pending.nation,
            dest_ids=list(pending.dest_ids),
            file_id=selected.file_id,
            file_name=selected.file_name,
        ), None

    @staticmethod
    def _extract_file_id(file_info: Dict[str, Any]) -> str:
        value = file_info.get("file_id") or file_info.get("fileId")
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _extract_file_name(file_info: Dict[str, Any]) -> str:
        raw = (
            file_info.get("file_name")
            or file_info.get("fileName")
            or file_info.get("name")
            or file_info.get("filename")
            or ""
        )
        name = str(raw).strip()
        return name if name else "未命名文件"

    @staticmethod
    def _build_selection_message(shipment_sn: str, options: List[_SelectableFileOption]) -> str:
        lines = [f"{shipment_sn} 未找到以 FBA 开头且 .pdf 结尾的文件。", "可选文件如下："]
        for idx, item in enumerate(options, start=1):
            lines.append(f"{idx}. {item.file_name}")
        lines.append("")
        lines.append("请直接回复文件名")
        return "\n".join(lines)

    @staticmethod
    def _build_response_from_selected_file(choice: _SelectedFileChoice) -> Dict[str, Any]:
        base_name = Path(choice.file_name).name or "selected.pdf"
        if not base_name.lower().endswith(".pdf"):
            base_name = f"{base_name}.pdf"
        items = []
        for dest_id in choice.dest_ids:
            item: Dict[str, Any] = {"destination_fulfillment_center_id": dest_id}
            if choice.nation is not None:
                item["nation"] = choice.nation
            items.append(item)
        if not items:
            item = {}
            if choice.nation is not None:
                item["nation"] = choice.nation
            items = [item]
        return {
            "data": [
                {
                    "shipment_sn": choice.shipment_sn,
                    "items": items,
                    "fileList": [
                        {
                            "file_id": choice.file_id,
                            "file_name": f"FBA_MANUAL_{base_name}",
                        }
                    ],
                }
            ]
        }

    def _log_request_event(
        self,
        *,
        request_id: str,
        message_id: Optional[str],
        user_id: Optional[str],
        user_name: Optional[str] = None,
        event_type: str,
        ack_status: Optional[str] = None,
        shipment_sns: Optional[List[str]] = None,
        detail: Optional[str] = None,
    ) -> None:
        if str(event_type).strip().upper() != "RECEIVED":
            return
        try:
            self._state_store.log_request_event(
                request_id=request_id,
                message_id=message_id,
                user_id=user_id,
                user_name=user_name,
                event_type=event_type,
                ack_status=ack_status,
                shipment_sns=shipment_sns,
                detail=detail,
            )
        except Exception as exc:
            self.logger.warning("[req=%s] Request log failed (%s): %s", request_id, event_type, exc)
            self._notify_tech_exception_sync(
                request_id=request_id,
                stage=f"request_log_{event_type}",
                exc=exc,
                shipment_sns=shipment_sns,
            )

    async def _worker_loop(self, worker_id: int) -> None:
        while True:
            job = await self._job_queue.get()
            try:
                await self._process_job(job, worker_id)
            except Exception as exc:
                self.logger.error("[req=%s] Worker %s crashed while processing: %s", job.request_id, worker_id, exc, exc_info=True)
                self._log_request_event(
                    request_id=job.request_id,
                    message_id=job.message_id,
                    user_id=job.sender_user_id,
                    event_type="PROCESS_CRASH",
                    ack_status="ERROR",
                    shipment_sns=job.shipment_sns,
                    detail=f"worker={worker_id}; error={exc}",
                )
                await self._notify_tech_exception(
                    request_id=job.request_id,
                    stage=f"worker_crash_{worker_id}",
                    exc=exc,
                    shipment_sns=job.shipment_sns,
                )
            finally:
                self._release_shipment_locks(job.shipment_sns, job.request_id)
                self._job_queue.task_done()

    async def _process_job(self, job: _QueuedJob, worker_id: int) -> None:
        incoming_message = dingtalk_stream.ChatbotMessage.from_dict(job.raw_data)
        sender_user_id = job.sender_user_id or _extract_sender_user_id(incoming_message)
        self._log_request_event(
            request_id=job.request_id,
            message_id=job.message_id,
            user_id=sender_user_id,
            event_type="PROCESS_START",
            ack_status="RUNNING",
            shipment_sns=job.shipment_sns,
            detail=f"worker={worker_id}",
        )
        self.logger.info(
            "[req=%s][worker=%s][user=%s] Start processing shipments=%s",
            job.request_id,
            worker_id,
            sender_user_id or "-",
            ",".join(job.shipment_sns),
        )

        await asyncio.to_thread(self._cleanup_downloads_if_needed, job.request_id)

        if job.selected_file is not None:
            response = self._build_response_from_selected_file(job.selected_file)
            _log_json(
                self.logger,
                "Manual file selection summary",
                {
                    "shipment_sn": job.selected_file.shipment_sn,
                    "file_name": job.selected_file.file_name,
                    "file_id": job.selected_file.file_id,
                },
                request_id=job.request_id,
            )
        else:
            try:
                response = await self._retry_async(
                    "LingXing fetch shipment detail",
                    lambda: self.client.fetch_shipment_list_detail(job.shipment_sns),
                    job.request_id,
                )
                if _is_param_error(response):
                    response = await self._retry_async(
                        "LingXing fetch shipment detail fallback",
                        lambda: self.client._request(  # type: ignore[attr-defined]
                            self.client.ROUTE_SHIPMENT_LIST_DETAIL,
                            req_body={"shipment_sn_arr": job.shipment_sns},
                        ),
                        job.request_id,
                    )
                _log_json(self.logger, "LingXing summary", _summarize_lingxing_response(response), request_id=job.request_id)
            except Exception as exc:
                self.logger.error("[req=%s] LingXing query failed: %s", job.request_id, exc, exc_info=True)
                self._log_request_event(
                    request_id=job.request_id,
                    message_id=job.message_id,
                    user_id=sender_user_id,
                    event_type="PROCESS_FAILED",
                    ack_status="ERROR",
                    shipment_sns=job.shipment_sns,
                    detail=f"stage=lingxing_query; error={exc}",
                )
                await self._notify_tech_exception(
                    request_id=job.request_id,
                    stage="lingxing_query",
                    exc=exc,
                    shipment_sns=job.shipment_sns,
                )
                await self._send_reply(
                    f"领星查询失败: {exc}",
                    incoming_message,
                    job.raw_data,
                    sender_user_id,
                    request_id=job.request_id,
                )
                return

        try:
            reply_result = await asyncio.to_thread(self._build_reply_with_ocr, response, job.shipment_sns, job.request_id)
        except Exception as exc:
            self.logger.error("[req=%s] Build reply failed: %s", job.request_id, exc, exc_info=True)
            self._log_request_event(
                request_id=job.request_id,
                message_id=job.message_id,
                user_id=sender_user_id,
                event_type="PROCESS_FAILED",
                ack_status="ERROR",
                shipment_sns=job.shipment_sns,
                detail=f"stage=build_reply; error={exc}",
            )
            await self._notify_tech_exception(
                request_id=job.request_id,
                stage="build_reply_with_ocr",
                exc=exc,
                shipment_sns=job.shipment_sns,
            )
            await self._send_reply(
                f"处理失败: {exc}",
                incoming_message,
                job.raw_data,
                sender_user_id,
                request_id=job.request_id,
            )
            return

        if reply_result.summary_text:
            await self._send_reply(
                reply_result.summary_text,
                incoming_message,
                job.raw_data,
                sender_user_id,
                request_id=job.request_id,
            )
        for message_text in reply_result.shipment_messages:
            await self._send_reply(
                message_text,
                incoming_message,
                job.raw_data,
                sender_user_id,
                request_id=job.request_id,
            )
        if sender_user_id:
            for pending in reply_result.pending_file_selections:
                self._register_pending_file_selection(sender_user_id, pending)
        for message_text in reply_result.selection_messages:
            await self._send_reply(
                message_text,
                incoming_message,
                job.raw_data,
                sender_user_id,
                request_id=job.request_id,
            )
        self._log_request_event(
            request_id=job.request_id,
            message_id=job.message_id,
            user_id=sender_user_id,
            event_type="PROCESS_SUCCESS",
            ack_status="OK",
            shipment_sns=job.shipment_sns,
            detail=f"worker={worker_id}",
        )
        self.logger.info("[req=%s] Job finished.", job.request_id)

    def _cleanup_downloads_if_needed(self, request_id: str) -> None:
        now_mono = time.monotonic()
        with self._cleanup_lock:
            if now_mono - self._last_cleanup_monotonic < self._download_cleanup_interval_sec:
                return
            self._last_cleanup_monotonic = now_mono

        cutoff_ts = time.time() - self._download_retention_days * 86400
        download_root = Path(config.DOWNLOAD_DIR)
        if not download_root.exists():
            return

        removed = 0
        for child in download_root.iterdir():
            try:
                stat = child.stat()
            except Exception:
                continue
            if stat.st_mtime >= cutoff_ts:
                continue
            try:
                if child.is_dir():
                    shutil.rmtree(child, ignore_errors=True)
                else:
                    child.unlink(missing_ok=True)
                removed += 1
            except Exception as exc:
                self.logger.warning("[req=%s] Cleanup failed for %s: %s", request_id, child, exc)
        if removed:
            self.logger.info("[req=%s] Cleanup removed %s expired download item(s).", request_id, removed)

    @staticmethod
    def _is_retryable_exception(exc: Exception) -> bool:
        if isinstance(exc, (TimeoutError, ConnectionError, urllib.error.URLError, asyncio.TimeoutError)):
            return True
        text = str(exc).lower()
        markers = (
            "timeout",
            "timed out",
            "temporarily",
            "temporary",
            "connection reset",
            "connection aborted",
            "connection refused",
            "service unavailable",
            "rate limit",
            "too many requests",
            " 429",
            " 500",
            " 502",
            " 503",
            " 504",
        )
        return any(marker in text for marker in markers)

    def _retry_sync(self, op_name: str, fn: Callable[[], Any], request_id: str) -> Any:
        for attempt in range(1, self._retry_times + 1):
            try:
                return fn()
            except Exception as exc:
                retryable = self._is_retryable_exception(exc)
                if attempt >= self._retry_times or not retryable:
                    raise
                delay = min(self._retry_max_delay_sec, self._retry_base_delay_sec * (2 ** (attempt - 1)))
                delay += random.uniform(0.0, 0.3)
                self.logger.warning(
                    "[req=%s] %s failed (%s/%s): %s. Retrying in %.2fs.",
                    request_id,
                    op_name,
                    attempt,
                    self._retry_times,
                    exc,
                    delay,
                )
                time.sleep(delay)
        raise RuntimeError(f"{op_name} retry loop exited unexpectedly")

    async def _retry_async(self, op_name: str, fn: Callable[[], Awaitable[Any]], request_id: str) -> Any:
        for attempt in range(1, self._retry_times + 1):
            try:
                return await fn()
            except Exception as exc:
                retryable = self._is_retryable_exception(exc)
                if attempt >= self._retry_times or not retryable:
                    raise
                delay = min(self._retry_max_delay_sec, self._retry_base_delay_sec * (2 ** (attempt - 1)))
                delay += random.uniform(0.0, 0.3)
                self.logger.warning(
                    "[req=%s] %s failed (%s/%s): %s. Retrying in %.2fs.",
                    request_id,
                    op_name,
                    attempt,
                    self._retry_times,
                    exc,
                    delay,
                )
                await asyncio.sleep(delay)
        raise RuntimeError(f"{op_name} retry loop exited unexpectedly")

    async def process(self, callback: dingtalk_stream.CallbackMessage) -> Tuple[str, str]:
        raw_data = callback.data
        request_id = uuid.uuid4().hex[:12]
        message_id: Optional[str] = None
        sender_user_id: Optional[str] = None
        shipment_sns: List[str] = []
        incoming_message = None

        try:
            await self._ensure_workers()
            incoming_message = dingtalk_stream.ChatbotMessage.from_dict(raw_data)
            msg_type = incoming_message.message_type
            sender_user_id = _extract_sender_user_id(incoming_message)
            sender_user_name = _extract_sender_user_name(incoming_message, raw_data)
            text_content = _extract_text_content(incoming_message)
            message_id = _extract_message_id(incoming_message, raw_data)

            if message_id:
                request_id = str(message_id).strip() or request_id
            if len(request_id) > 64:
                request_id = request_id[:64]

            self._log_request_event(
                request_id=request_id,
                message_id=message_id,
                user_id=sender_user_id,
                user_name=sender_user_name,
                event_type="RECEIVED",
                ack_status=None,
                shipment_sns=None,
                detail=(text_content or f"[{msg_type}]"),
            )

            if message_id:
                is_new = self._register_message_once(message_id, request_id=request_id)
                if not is_new:
                    self.logger.info("[req=%s] Duplicate message ignored: %s", request_id, message_id)
                    self._log_request_event(
                        request_id=request_id,
                        message_id=message_id,
                        user_id=sender_user_id,
                        event_type="DEDUP_REJECT",
                        ack_status="DUPLICATE",
                        shipment_sns=None,
                        detail="duplicate message ignored",
                    )
                    await self._send_reply(
                        "重复消息已忽略，请勿重复提交。",
                        incoming_message,
                        raw_data,
                        sender_user_id,
                        request_id=request_id,
                    )
                    return AckMessage.STATUS_OK, "DUPLICATE"

            if msg_type != "text":
                self._log_request_event(
                    request_id=request_id,
                    message_id=message_id,
                    user_id=sender_user_id,
                    event_type="INVALID_REQUEST",
                    ack_status="OK",
                    shipment_sns=None,
                    detail=f"unsupported message_type={msg_type}",
                )
                await self._send_reply("Only text messages are supported.", incoming_message, raw_data, sender_user_id, request_id=request_id)
                return AckMessage.STATUS_OK, "OK"

            if not text_content:
                self._log_request_event(
                    request_id=request_id,
                    message_id=message_id,
                    user_id=sender_user_id,
                    event_type="INVALID_REQUEST",
                    ack_status="OK",
                    shipment_sns=None,
                    detail="empty text content",
                )
                await self._send_reply("Please send shipment numbers like SP260119001.", incoming_message, raw_data, sender_user_id, request_id=request_id)
                return AckMessage.STATUS_OK, "OK"

            parsed_selection = _parse_file_selection_command(text_content)
            if parsed_selection:
                shipment_sn, selection_index = parsed_selection
                if not sender_user_id:
                    await self._send_reply(
                        "无法识别你的用户信息，请重新发送发货单号后再试。",
                        incoming_message,
                        raw_data,
                        sender_user_id,
                        request_id=request_id,
                    )
                    return AckMessage.STATUS_OK, "OK"
                selected_file, selection_error = self._consume_pending_file_selection(
                    sender_user_id,
                    shipment_sn,
                    selection_index,
                )
                if selection_error or selected_file is None:
                    await self._send_reply(
                        selection_error or "未找到待选择文件，请重新发送发货单号。",
                        incoming_message,
                        raw_data,
                        sender_user_id,
                        request_id=request_id,
                    )
                    return AckMessage.STATUS_OK, "OK"

                lock_ok, busy_sn = self._acquire_shipment_locks([selected_file.shipment_sn], request_id)
                if not lock_ok:
                    await self._send_reply(
                        f"发货单 {busy_sn or selected_file.shipment_sn} 正在处理中，请稍后重试。",
                        incoming_message,
                        raw_data,
                        sender_user_id,
                        request_id=request_id,
                    )
                    return AckMessage.STATUS_OK, "BUSY"

                job = _QueuedJob(
                    request_id=request_id,
                    message_id=message_id,
                    raw_data=raw_data,
                    sender_user_id=sender_user_id,
                    shipment_sns=[selected_file.shipment_sn],
                    queued_at=time.time(),
                    selected_file=selected_file,
                )
                try:
                    self._job_queue.put_nowait(job)
                except asyncio.QueueFull:
                    self._release_shipment_locks([selected_file.shipment_sn], request_id)
                    await self._send_reply(
                        "当前排队任务过多，请稍后再试。",
                        incoming_message,
                        raw_data,
                        sender_user_id,
                        request_id=request_id,
                    )
                    return AckMessage.STATUS_OK, "QUEUE_FULL"

                queue_size = self._job_queue.qsize()
                ahead_count = max(queue_size - 1, 0)
                queue_reply = "已接收你选择的文件，正在核对..." if ahead_count == 0 else f"已接收你选择的文件，前面{ahead_count}人，请稍等..."
                await self._send_reply(
                    queue_reply,
                    incoming_message,
                    raw_data,
                    sender_user_id,
                    request_id=request_id,
                )
                return AckMessage.STATUS_OK, "QUEUED"
            elif sender_user_id:
                selected_file, selection_error = self._consume_pending_file_selection_by_name(
                    sender_user_id,
                    text_content,
                )
                if selection_error:
                    await self._send_reply(
                        selection_error,
                        incoming_message,
                        raw_data,
                        sender_user_id,
                        request_id=request_id,
                    )
                    return AckMessage.STATUS_OK, "OK"
                if selected_file is not None:
                    lock_ok, busy_sn = self._acquire_shipment_locks([selected_file.shipment_sn], request_id)
                    if not lock_ok:
                        await self._send_reply(
                            f"发货单 {busy_sn or selected_file.shipment_sn} 正在处理中，请稍后重试。",
                            incoming_message,
                            raw_data,
                            sender_user_id,
                            request_id=request_id,
                        )
                        return AckMessage.STATUS_OK, "BUSY"

                    job = _QueuedJob(
                        request_id=request_id,
                        message_id=message_id,
                        raw_data=raw_data,
                        sender_user_id=sender_user_id,
                        shipment_sns=[selected_file.shipment_sn],
                        queued_at=time.time(),
                        selected_file=selected_file,
                    )
                    try:
                        self._job_queue.put_nowait(job)
                    except asyncio.QueueFull:
                        self._release_shipment_locks([selected_file.shipment_sn], request_id)
                        await self._send_reply(
                            "当前排队任务过多，请稍后再试。",
                            incoming_message,
                            raw_data,
                            sender_user_id,
                            request_id=request_id,
                        )
                        return AckMessage.STATUS_OK, "QUEUE_FULL"

                    queue_size = self._job_queue.qsize()
                    ahead_count = max(queue_size - 1, 0)
                    queue_reply = "已接收你选择的文件，正在核对..." if ahead_count == 0 else f"已接收你选择的文件，前面{ahead_count}人，请稍等..."
                    await self._send_reply(
                        queue_reply,
                        incoming_message,
                        raw_data,
                        sender_user_id,
                        request_id=request_id,
                    )
                    return AckMessage.STATUS_OK, "QUEUED"

            shipment_sns = _extract_shipment_sns(text_content)
            if not shipment_sns:
                self._log_request_event(
                    request_id=request_id,
                    message_id=message_id,
                    user_id=sender_user_id,
                    event_type="INVALID_REQUEST",
                    ack_status="OK",
                    shipment_sns=None,
                    detail="shipment number not found",
                )
                await self._send_reply("No shipment number found. Example: SP260119001", incoming_message, raw_data, sender_user_id, request_id=request_id)
                return AckMessage.STATUS_OK, "OK"

            if not config.LINGXING_API_KEY or not config.LINGXING_API_SECRET:
                self._log_request_event(
                    request_id=request_id,
                    message_id=message_id,
                    user_id=sender_user_id,
                    event_type="CONFIG_ERROR",
                    ack_status="OK",
                    shipment_sns=shipment_sns,
                    detail="missing LingXing API credentials",
                )
                await self._send_reply("Missing LingXing API credentials.", incoming_message, raw_data, sender_user_id, request_id=request_id)
                return AckMessage.STATUS_OK, "OK"

            lock_ok, busy_sn = self._acquire_shipment_locks(shipment_sns, request_id)
            if not lock_ok:
                raise_busy = ShipmentBusyError(busy_sn or "unknown")
                self._log_request_event(
                    request_id=request_id,
                    message_id=message_id,
                    user_id=sender_user_id,
                    event_type="LOCK_BUSY",
                    ack_status="BUSY",
                    shipment_sns=shipment_sns,
                    detail=f"busy_shipment={busy_sn or 'unknown'}",
                )
                await self._send_reply(
                    f"发货单 {raise_busy} 正在处理中，请稍后重试。",
                    incoming_message,
                    raw_data,
                    sender_user_id,
                    request_id=request_id,
                )
                return AckMessage.STATUS_OK, "BUSY"

            job = _QueuedJob(
                request_id=request_id,
                message_id=message_id,
                raw_data=raw_data,
                sender_user_id=sender_user_id,
                shipment_sns=shipment_sns,
                queued_at=time.time(),
            )

            try:
                self._job_queue.put_nowait(job)
            except asyncio.QueueFull:
                self._release_shipment_locks(shipment_sns, request_id)
                self._log_request_event(
                    request_id=request_id,
                    message_id=message_id,
                    user_id=sender_user_id,
                    event_type="QUEUE_REJECT",
                    ack_status="QUEUE_FULL",
                    shipment_sns=shipment_sns,
                    detail=f"queue_max_size={self._queue_max_size}",
                )
                await self._send_reply(
                    "当前排队任务过多，请稍后再试。",
                    incoming_message,
                    raw_data,
                    sender_user_id,
                    request_id=request_id,
                )
                return AckMessage.STATUS_OK, "QUEUE_FULL"

            queue_size = self._job_queue.qsize()
            self._log_request_event(
                request_id=request_id,
                message_id=message_id,
                user_id=sender_user_id,
                event_type="QUEUED",
                ack_status="QUEUED",
                shipment_sns=shipment_sns,
                detail=f"queue_size={queue_size}",
            )
            ahead_count = max(queue_size - 1, 0)
            if ahead_count == 0:
                queue_reply = "已接收，正在核对..."
            else:
                queue_reply = f"已接收，前面{ahead_count}人，请稍等..."
            await self._send_reply(
                queue_reply,
                incoming_message,
                raw_data,
                sender_user_id,
                request_id=request_id,
            )
            return AckMessage.STATUS_OK, "QUEUED"
        except Exception as exc:
            self.logger.error("[req=%s] Callback processing failed: %s", request_id, exc, exc_info=True)
            await self._notify_tech_exception(
                request_id=request_id,
                stage="callback_process",
                exc=exc,
                shipment_sns=shipment_sns or None,
            )
            if incoming_message is not None:
                try:
                    await self._send_reply(
                        "系统异常，请联系技术人员。",
                        incoming_message,
                        raw_data,
                        sender_user_id,
                        request_id=request_id,
                    )
                except Exception:
                    pass
            return AckMessage.STATUS_OK, "ERROR"

    async def _send_reply(
        self,
        text: str,
        incoming_message,
        raw_data,
        user_id: Optional[str],
        attachment_files: Optional[List[Path]] = None,
        request_id: Optional[str] = None,
    ) -> None:
        text_sent = False
        session_webhook = _extract_session_webhook(incoming_message, raw_data)
        if session_webhook:
            try:
                await asyncio.to_thread(_send_session_text, session_webhook, text)
                text_sent = True
            except Exception as exc:
                self.logger.warning("Session webhook send failed: %s", exc)
                await self._notify_tech_exception(
                    request_id=request_id,
                    stage="send_reply_session_webhook",
                    exc=exc,
                )

        if not text_sent and not user_id:
            self.logger.warning("Missing user_id; cannot send fallback reply.")
            return
        if not text_sent and user_id:
            try:
                await asyncio.to_thread(_DEFAULT_NOTIFIER.send_user_text, user_id, text)
            except Exception as exc:
                self.logger.error("[req=%s] Fallback user reply failed(user=%s): %s", request_id or "-", user_id, exc)
                await self._notify_tech_exception(
                    request_id=request_id,
                    stage="send_reply_user_text",
                    exc=exc,
                )
                return

        if not attachment_files:
            return
        if not user_id:
            self.logger.warning("Missing user_id; cannot send image attachments.")
            return

        for file_path in attachment_files:
            try:
                path = Path(file_path)
                if not path.exists():
                    self.logger.warning("Attachment file does not exist: %s", path)
                    continue
                await asyncio.to_thread(_DEFAULT_NOTIFIER.send_user_file, user_id, str(path))
            except Exception as exc:
                self.logger.warning("Send attachment failed (%s): %s", file_path, exc)
                await self._notify_tech_exception(
                    request_id=request_id,
                    stage="send_reply_attachment",
                    exc=exc,
                )

    @staticmethod
    def _normalize_data_list(data: Any) -> List[Dict[str, Any]]:
        if isinstance(data, list):
            return [item for item in data if isinstance(item, dict)]
        if isinstance(data, dict):
            for key in ("data", "list", "rows", "items"):
                value = data.get(key)
                if isinstance(value, list):
                    return [item for item in value if isinstance(item, dict)]
            nested = data.get("data") or data.get("result")
            if isinstance(nested, dict):
                for key in ("data", "list", "rows", "items"):
                    value = nested.get(key)
                    if isinstance(value, list):
                        return [item for item in value if isinstance(item, dict)]
            return [data]
        return []

    def _get_ocr_client(self) -> AliyunOCRClient:
        if self._ocr_client is not None:
            return self._ocr_client
        with self._ocr_client_lock:
            if self._ocr_client is None:
                self._ocr_client = AliyunOCRClient()
        return self._ocr_client

    def _load_address_book_if_needed(self) -> None:
        with self._address_book_lock:
            # Reload on every call so Excel updates take effect immediately.
            self._address_book_loaded = False
            self._address_book_disabled_reason = None
            self._address_book_by_country_port = {}

            path = Path(config.ADDRESS_BOOK_XLSX_PATH)
            if path.is_dir():
                path = path / "全站点地址.xlsx"
            if not path.exists():
                self._address_book_disabled_reason = f"地址簿不存在: {path}"
                self._address_book_loaded = True
                self.logger.warning(self._address_book_disabled_reason)
                return

            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception as exc:
                self._address_book_disabled_reason = f"openpyxl 不可用: {exc}"
                self._address_book_loaded = True
                self.logger.warning(self._address_book_disabled_reason)
                return

            workbook = None
            loaded_country_count = 0
            try:
                workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
                sheet_lookup = {str(name).strip().lower(): name for name in workbook.sheetnames}
                country_sheet_candidates = {
                    "US": ("美国",),
                    "DE": ("德国",),
                    "UK": ("英国",),
                    "AU": ("澳洲", "澳大利亚"),
                }

                for country_mode, candidates in country_sheet_candidates.items():
                    sheet_name = None
                    for candidate in candidates:
                        key = str(candidate).strip().lower()
                        if key in sheet_lookup:
                            sheet_name = sheet_lookup[key]
                            break
                    if not sheet_name:
                        continue

                    sheet = workbook[sheet_name]
                    header_rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
                    header_row = next(header_rows, None)
                    if not header_row:
                        continue

                    header_map: Dict[str, int] = {}
                    for idx, value in enumerate(header_row, start=1):
                        key = str(value).strip() if value is not None else ""
                        if key:
                            header_map[key] = idx
                    receiver_col = header_map.get("收件人")
                    port_col = header_map.get("目的港")
                    if not receiver_col or not port_col:
                        self.logger.warning(
                            "Address book sheet %s missing columns 收件人/目的港, skip address matching for %s.",
                            sheet_name,
                            country_mode,
                        )
                        continue

                    port_mapping: Dict[str, List[_AddressRecord]] = {}
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if not row:
                            continue
                        receiver_val = row[receiver_col - 1] if len(row) >= receiver_col else None
                        port_val = row[port_col - 1] if len(row) >= port_col else None
                        receiver_text = str(receiver_val).strip() if receiver_val is not None else ""
                        port_text = str(port_val).strip() if port_val is not None else ""
                        port_code = _normalize_fc_code(port_text)
                        if not receiver_text or not port_code:
                            continue
                        parsed = _parse_address_record(receiver_text, country_mode, fc_code=port_code)
                        if not _has_required_fields(parsed, country_mode):
                            continue
                        port_mapping.setdefault(port_code, []).append(parsed)

                    if port_mapping:
                        self._address_book_by_country_port[country_mode] = port_mapping
                        loaded_country_count += 1
                        self.logger.info(
                            "Address book loaded for %s from sheet %s with %s destination codes.",
                            country_mode,
                            sheet_name,
                            len(port_mapping),
                        )

            except Exception as exc:
                self._address_book_disabled_reason = f"加载地址簿失败: {exc}"
                self.logger.warning(self._address_book_disabled_reason)
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        pass
                self._address_book_loaded = True

            if loaded_country_count == 0 and not self._address_book_disabled_reason:
                self._address_book_disabled_reason = "地址簿未加载到美国/德国/英国/澳洲可用数据。"
                self.logger.warning(self._address_book_disabled_reason)

    def _validate_address_fallback(
        self,
        *,
        request_id: Optional[str],
        shipment_sn: str,
        nation: Optional[str],
        dest_ids: List[str],
        ocr_text: str,
    ) -> Tuple[bool, str, Optional[_AddressMatchOutcome]]:
        country_mode = _detect_country_mode_by_nation(nation)
        if country_mode not in SUPPORTED_ADDRESS_COUNTRIES:
            return False, f"ADDR_UNSUPPORTED_COUNTRY: nation={nation or '-'}", None

        self.logger.info(
            "[req=%s][sn=%s] Address matching started. nation=%s country_mode=%s dest_ids=%s",
            request_id or "-",
            shipment_sn,
            nation or "-",
            country_mode,
            ",".join(dest_ids) or "-",
        )

        self._load_address_book_if_needed()
        if self._address_book_disabled_reason:
            return False, f"ADDR_PARSE_INCOMPLETE: {self._address_book_disabled_reason}", None

        country_mapping = self._address_book_by_country_port.get(country_mode)
        if not country_mapping:
            return False, f"ADDR_PARSE_INCOMPLETE: 地址簿缺少国家sheet数据({nation or country_mode})", None

        candidate_records: List[_AddressRecord] = []
        checked_ports: List[str] = []
        for dest_id in _unique_preserve_order(dest_ids):
            port_code = _normalize_fc_code(dest_id)
            if not port_code:
                continue
            checked_ports.append(port_code)
            candidate_records.extend(country_mapping.get(port_code) or [])

        if not candidate_records:
            return False, (
                f"ADDR_PARSE_INCOMPLETE: 地址簿未找到候选地址(country={nation or country_mode}, "
                f"ports={','.join(checked_ports) or '-'})"
            ), None

        destination_text = _extract_destination_block(ocr_text)
        parsed_ocr = _parse_address_record(destination_text or ocr_text, country_mode)
        if not _has_required_fields(parsed_ocr, country_mode):
            parsed_ocr = _parse_address_record(ocr_text, country_mode)
        missing_fields = _missing_required_fields(parsed_ocr, country_mode)
        if missing_fields:
            return False, f"ADDR_PARSE_INCOMPLETE: missing={','.join(missing_fields)}", None

        self.logger.info(
            "[req=%s][sn=%s] Parsed OCR address: %s",
            request_id or "-",
            shipment_sn,
            _format_address_record_for_reply(parsed_ocr),
        )

        candidate_scores = [
            _score_address_candidate(country_mode, parsed_ocr, candidate)
            for candidate in candidate_records
        ]
        hard_passed = [item for item in candidate_scores if item.hard_ok]
        candidate_scores.sort(key=lambda item: item.score, reverse=True)
        hard_passed.sort(key=lambda item: item.score, reverse=True)
        if candidate_scores:
            top = candidate_scores[0]
            self.logger.info(
                "[req=%s][sn=%s] Top candidate before hard-filter: score=%.1f hard_ok=%s reason=%s candidate=%s",
                request_id or "-",
                shipment_sn,
                top.score,
                top.hard_ok,
                top.hard_reason or "-",
                _format_address_record_for_reply(top.record),
            )

        if not hard_passed:
            top = candidate_scores[0]
            return False, (
                f"ADDR_HARD_MISMATCH: reason={top.hard_reason}; top_score={top.score:.1f}; "
                f"fc={top.record.fc_code}"
            ), None

        top = hard_passed[0]
        second = hard_passed[1] if len(hard_passed) > 1 else None
        gap = (top.score - second.score) if second else 100.0
        if top.score < ADDRESS_MATCH_PASS_SCORE:
            return False, (
                f"ADDR_LOW_SCORE: top={top.score:.1f}; pass={ADDRESS_MATCH_PASS_SCORE:.1f}; "
                f"fc={top.record.fc_code}"
            ), None
        if second and gap < ADDRESS_MATCH_MIN_GAP:
            return False, (
                f"ADDR_AMBIGUOUS: top={top.score:.1f}; second={second.score:.1f}; "
                f"gap={gap:.1f}; min_gap={ADDRESS_MATCH_MIN_GAP:.1f}"
            ), None

        outcome = _AddressMatchOutcome(
            fc_code=top.record.fc_code,
            score=top.score,
            gap=gap,
            postal_code=top.record.postal_code or top.record.zip5,
            city=top.record.city,
            ocr_extracted=_format_ocr_record_for_reply(parsed_ocr),
            excel_extracted=str(top.record.raw_text or "").strip(),
        )
        self.logger.info(
            "[req=%s][sn=%s] Address matching passed. score=%.1f gap=%.1f fc=%s city=%s postal=%s",
            request_id or "-",
            shipment_sn,
            outcome.score,
            outcome.gap,
            outcome.fc_code or "-",
            outcome.city or "-",
            outcome.postal_code or "-",
        )
        return True, (
            f"ADDR_MATCH_OK: score={outcome.score:.1f}; gap={outcome.gap:.1f}; "
            f"fc={outcome.fc_code or '-'}"
        ), outcome

    def _build_reply_with_ocr(
        self,
        response: Dict[str, Any],
        requested_sns: List[str],
        request_id: Optional[str] = None,
    ) -> _ReplyResult:
        data = response.get("data")
        data_list = self._normalize_data_list(data)
        if not data_list:
            code = response.get("code") or response.get("Code") or response.get("status")
            msg = response.get("message") or response.get("msg") or response.get("Message") or ""
            if code or msg:
                return _ReplyResult(
                    summary_text=f"No shipment data found. code={code} message={msg}",
                    shipment_messages=[],
                    shipment_images=[],
                    selection_messages=[],
                    pending_file_selections=[],
                )
            return _ReplyResult(
                summary_text="No shipment data found.",
                shipment_messages=[],
                shipment_images=[],
                selection_messages=[],
                pending_file_selections=[],
            )

        issues: List[str] = []
        shipment_messages: List[str] = []
        shipment_images: List[Path] = []
        selection_messages: List[str] = []
        pending_file_selections: List[_PendingFileSelection] = []
        found_sns = set()

        download_root = Path(config.DOWNLOAD_DIR)
        download_root.mkdir(parents=True, exist_ok=True)

        try:
            ocr_client = self._get_ocr_client()
        except Exception as exc:
            return _ReplyResult(
                summary_text=f"OCR client init failed: {exc}",
                shipment_messages=[],
                shipment_images=[],
                selection_messages=[],
                pending_file_selections=[],
            )

        run_dir = download_root / f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{uuid.uuid4().hex[:8]}"
        run_dir.mkdir(parents=True, exist_ok=True)
        for shipment in data_list:
            shipment_sn = (
                shipment.get("shipment_sn")
                or shipment.get("shipment_no")
                or shipment.get("shipmentNo")
                or ""
            )
            shipment_sn = str(shipment_sn).strip()
            if shipment_sn:
                found_sns.add(shipment_sn)

            items = shipment.get("items") or []
            nation = _extract_first_item_nation(items)
            dest_ids = []
            if isinstance(items, list):
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    dest = item.get("destination_fulfillment_center_id")
                    if dest:
                        dest_ids.append(str(dest).strip().upper())
            dest_ids = _unique_preserve_order([d for d in dest_ids if d])

            file_list = shipment.get("fileList") or shipment.get("file_list") or []
            pdf_files = []
            if isinstance(file_list, list):
                for file_info in file_list:
                    if not isinstance(file_info, dict):
                        continue
                    name = self._extract_file_name(file_info)
                    if name.upper().startswith("FBA") and name.lower().endswith(".pdf"):
                        pdf_files.append(file_info)

            if not shipment_sn:
                issues.append("未识别到发货单号")
                continue

            if len(pdf_files) == 0:
                selectable_options: List[_SelectableFileOption] = []
                if isinstance(file_list, list):
                    for file_info in file_list:
                        if not isinstance(file_info, dict):
                            continue
                        file_id = self._extract_file_id(file_info)
                        if not file_id:
                            continue
                        selectable_options.append(
                            _SelectableFileOption(
                                file_id=file_id,
                                file_name=self._extract_file_name(file_info),
                            )
                        )
                if selectable_options:
                    selection_messages.append(self._build_selection_message(shipment_sn, selectable_options))
                    pending_file_selections.append(
                        _PendingFileSelection(
                            shipment_sn=shipment_sn,
                            nation=nation,
                            dest_ids=list(dest_ids),
                            options=selectable_options,
                            expires_at=time.time() + self._pending_file_selection_ttl_sec,
                        )
                    )
                    issues.append(f"{shipment_sn} 未找到 FBA PDF，已发送可选文件列表，请按提示回复。")
                else:
                    issues.append(f"{shipment_sn} 未找到 FBA PDF")
                continue
            if len(pdf_files) > 1:
                issues.append(f"{shipment_sn} FBA PDF 数量异常: {len(pdf_files)}")
                continue

            pdf_info = pdf_files[0]
            file_id = self._extract_file_id(pdf_info)
            if not file_id:
                issues.append(f"{shipment_sn} PDF 缺少 file_id")
                continue

            safe_name = Path(self._extract_file_name(pdf_info) or "shipment.pdf").name
            shipment_dir = run_dir / (shipment_sn or "unknown")
            shipment_dir.mkdir(parents=True, exist_ok=True)
            pdf_path = shipment_dir / safe_name

            try:
                with self._acquire_limiter(self._download_limiter, "download"):
                    download_resp = self._retry_sync(
                        "LingXing download_common_file",
                        lambda: asyncio.run(self.client.download_common_file(file_id)),
                        request_id or "n/a",
                    )
                url, content, file_name = _extract_download_payload(download_resp)
                if file_name:
                    pdf_path = shipment_dir / Path(str(file_name)).name
                if content:
                    _save_pdf_bytes(content, pdf_path)
                elif url:
                    self._retry_sync(
                        "HTTP PDF download",
                        lambda: _download_file(str(url), pdf_path, expect_pdf=True),
                        request_id or "n/a",
                    )
                else:
                    detail_msg = (
                        download_resp.get("msg")
                        or download_resp.get("message")
                        or "download response missing url/content"
                    )
                    raise RuntimeError(str(detail_msg))
            except Exception as exc:
                issues.append(f"{shipment_sn} 下载失败: {exc}")
                continue

            image_path = shipment_dir / f"{shipment_sn}.png"
            try:
                _pdf_first_page_to_png(pdf_path, image_path)
            except Exception as exc:
                issues.append(f"{shipment_sn} PDF处理失败: {exc}")
                continue

            try:
                with self._acquire_limiter(self._ocr_limiter, "ocr"):
                    ocr_result = self._retry_sync(
                        "Aliyun OCR",
                        lambda: ocr_client.recognize_all_text(file_path=str(image_path), image_type="Advanced"),
                        request_id or "n/a",
                    )
            except Exception as exc:
                issues.append(f"{shipment_sn} OCR调用失败: {exc}")
                continue

            _log_json(self.logger, f"OCR result ({shipment_sn})", ocr_result, request_id=request_id)
            ocr_text = _extract_ocr_text(ocr_result)
            if not ocr_text.strip():
                issues.append(f"{shipment_sn} OCR结果为空")
                continue
            destination_block_text = _extract_destination_block(ocr_text)
            ocr_fc_candidates = _extract_fc_codes_filtered(destination_block_text)
            if not ocr_fc_candidates:
                ocr_fc_candidates = _extract_fc_codes_filtered(ocr_text)
            ocr_fc = ocr_fc_candidates[0] if ocr_fc_candidates else None
            ocr_fc_display = "/".join(ocr_fc_candidates) if ocr_fc_candidates else "-"

            if not dest_ids:
                issues.append(f"{shipment_sn} 未找到 destination_fulfillment_center_id")
                continue

            fc_match = bool(ocr_fc and ocr_fc in dest_ids)
            matched_address: Optional[_AddressMatchOutcome] = None
            match_basis = ""
            ocr_extracted_text = ""
            excel_extracted_text = ""
            country_mode = _detect_country_mode_by_nation(nation)
            if not fc_match:
                if _is_fc_contains_country(nation):
                    self.logger.info(
                        "[req=%s][sn=%s] FC-contains fallback started. nation=%s dest_ids=%s ocr_fc=%s",
                        request_id or "-",
                        shipment_sn,
                        nation or "-",
                        ",".join(dest_ids) or "-",
                        ocr_fc or "-",
                    )
                    if _dest_code_contains_in_ocr_text(dest_ids, ocr_text):
                        self.logger.info(
                            "[req=%s][sn=%s] FC-contains fallback matched. dest_ids=%s",
                            request_id or "-",
                            shipment_sn,
                            ",".join(dest_ids) or "-",
                        )
                        fc_match = True
                        if ocr_fc:
                            match_basis = f"FC文本包含匹配(OCR={ocr_fc}, 领星FC={','.join(dest_ids)})"
                        else:
                            match_basis = f"FC文本包含匹配(OCR未识别FC, 领星FC={','.join(dest_ids)})"
                        if country_mode in SUPPORTED_ADDRESS_COUNTRIES:
                            parsed_ocr_for_reply = _parse_address_record(
                                _extract_destination_block(ocr_text) or ocr_text,
                                country_mode,
                            )
                            parsed_ocr_for_reply.fc_code = _normalize_fc_code(ocr_fc or "")
                            ocr_extracted_text = _format_ocr_record_for_reply(parsed_ocr_for_reply)
                        else:
                            ocr_extracted_text = ""
                        excel_extracted_text = ""
                    else:
                        self.logger.info(
                            "[req=%s][sn=%s] FC-contains fallback not matched. dest_ids=%s",
                            request_id or "-",
                            shipment_sn,
                            ",".join(dest_ids) or "-",
                        )
                        if not ocr_fc:
                            issues.append(f"{shipment_sn} OCR 未识别到目的地，且OCR文本未包含物流中心编码({','.join(dest_ids)})")
                        else:
                            issues.append(
                                f"{shipment_sn} 目的地不一致: OCR候选={ocr_fc_display}, 领星={','.join(dest_ids)}，且OCR文本未包含物流中心编码"
                            )
                        continue
            else:
                match_basis = f"FC直匹配(OCR={ocr_fc}, 领星FC={','.join(dest_ids)})"
                if country_mode in SUPPORTED_ADDRESS_COUNTRIES:
                    parsed_ocr_for_reply = _parse_address_record(
                        _extract_destination_block(ocr_text) or ocr_text,
                        country_mode,
                    )
                    parsed_ocr_for_reply.fc_code = _normalize_fc_code(ocr_fc or "")
                    ocr_extracted_text = _format_ocr_record_for_reply(parsed_ocr_for_reply)
                else:
                    ocr_extracted_text = ""
                excel_extracted_text = ""

            if not fc_match:
                fallback_ok, fallback_detail, matched_address = self._validate_address_fallback(
                    request_id=request_id,
                    shipment_sn=shipment_sn,
                    nation=nation,
                    dest_ids=dest_ids,
                    ocr_text=ocr_text,
                )
                if not fallback_ok:
                    self.logger.info(
                        "[req=%s][sn=%s] Address matching rejected: %s",
                        request_id or "-",
                        shipment_sn,
                        fallback_detail,
                    )
                    if not ocr_fc:
                        issues.append(f"{shipment_sn} OCR 未识别到目的地; {fallback_detail}")
                    else:
                        issues.append(
                            f"{shipment_sn} 目的地不一致: OCR候选={ocr_fc_display}, 领星={','.join(dest_ids)}; {fallback_detail}"
                        )
                    continue
                if matched_address:
                    match_basis = (
                        "地址核对匹配("
                        f"score={matched_address.score:.1f}, gap={matched_address.gap:.1f}, "
                        f"fc={matched_address.fc_code or '-'}, city={matched_address.city or '-'}, "
                        f"postal={matched_address.postal_code or '-'}"
                        ")"
                    )
                    ocr_extracted_text = matched_address.ocr_extracted
                    excel_extracted_text = matched_address.excel_extracted
                else:
                    match_basis = f"地址核对匹配({fallback_detail})"

            logistics_code = dest_ids[0] if dest_ids else "未知"
            country_name = nation or "未知"
            shipment_message = (
                f"发货单号：{shipment_sn}\n"
                f"物流中心编码：{logistics_code}\n"
                f"国家：{country_name}"
            )
            if ocr_extracted_text:
                shipment_message += f"\n\n---------------- OCR提取 ----------------\n{ocr_extracted_text}"
            if excel_extracted_text:
                shipment_message += f"\n\n---------------- 全站点Excel内容 ----------------\n{excel_extracted_text}"
            shipment_messages.append(shipment_message)
            shipment_images.append(image_path)

        missing = [sn for sn in requested_sns if sn not in found_sns]
        if missing:
            issues.append(f"未找到发货单: {', '.join(missing)}")

        summary_text = None
        if issues:
            summary_text = "异常发货单:\n" + "\n".join(issues)
        elif shipment_messages:
            summary_text = f"校对完成：{len(shipment_messages)}个发货单一致。"
        else:
            summary_text = "没有可发送的发货单结果。"

        return _ReplyResult(
            summary_text=summary_text,
            shipment_messages=shipment_messages,
            shipment_images=shipment_images,
            selection_messages=selection_messages,
            pending_file_selections=pending_file_selections,
        )
