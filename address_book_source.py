#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Address book workbook loading from local files or SMB shares."""

from __future__ import annotations

import socket
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Optional, Tuple
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook  # type: ignore


@dataclass
class AddressBookWorkbook:
    workbook: Any
    source_label: str

    def close(self) -> None:
        close_fn = getattr(self.workbook, "close", None)
        if callable(close_fn):
            close_fn()


def _is_smb_path(path: str) -> bool:
    value = str(path or "").strip()
    return value.lower().startswith("smb://") or value.startswith("\\\\") or value.startswith("//")


def _parse_smb_path(path: str) -> Tuple[str, str, str]:
    value = str(path or "").strip()
    if value.lower().startswith("smb://"):
        parsed = urlparse(value)
        host = unquote(parsed.hostname or "")
        pieces = [unquote(part) for part in parsed.path.split("/") if part]
    elif value.startswith("\\\\"):
        parts = [part for part in value.strip("\\").split("\\") if part]
        host = parts[0] if parts else ""
        pieces = parts[1:]
    elif value.startswith("//"):
        parts = [unquote(part) for part in value.strip("/").split("/") if part]
        host = parts[0] if parts else ""
        pieces = parts[1:]
    else:
        raise ValueError(f"Not an SMB path: {path}")

    if len(pieces) < 2 or not host:
        raise ValueError(f"SMB path must include host, share, and file path: {path}")

    share = pieces[0]
    remote_path = "/" + "/".join(pieces[1:])
    return host, share, remote_path


def _read_smb_file(
    *,
    host: str,
    share: str,
    remote_path: str,
    username: str,
    password: str,
    port: int,
    timeout_sec: int,
    client_name: Optional[str],
) -> BytesIO:
    if not username:
        raise RuntimeError("SMB username is required for address book access")
    if not password:
        raise RuntimeError("SMB password is required for address book access")

    try:
        from smb.SMBConnection import SMBConnection  # type: ignore
    except Exception as exc:
        raise RuntimeError("pysmb is required for SMB address book access") from exc

    buffer = BytesIO()
    machine_name = client_name or socket.gethostname() or "dingtalk-cp-bot"
    conn = SMBConnection(
        username,
        password,
        machine_name,
        host,
        use_ntlm_v2=True,
        is_direct_tcp=True,
    )
    try:
        if not conn.connect(host, port, timeout=timeout_sec):
            raise RuntimeError(f"SMB connect/auth failed: {host}:{port}")
        conn.retrieveFile(share, remote_path, buffer, timeout=timeout_sec)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    buffer.seek(0)
    return buffer


def open_address_book_workbook(
    *,
    path: str,
    smb_host: str = "",
    smb_share: str = "",
    smb_username: str = "",
    smb_password: str = "",
    smb_port: int = 445,
    smb_timeout_sec: int = 30,
    smb_client_name: str = "",
) -> AddressBookWorkbook:
    """Open the address book workbook from a local path or SMB path."""
    raw_path = str(path or "").strip()
    if not raw_path:
        raise RuntimeError("ADDRESS_BOOK_XLSX_PATH is empty")

    if _is_smb_path(raw_path):
        host, share, remote_path = _parse_smb_path(raw_path)
        data = _read_smb_file(
            host=smb_host.strip() or host,
            share=smb_share.strip() or share,
            remote_path=remote_path,
            username=smb_username.strip(),
            password=smb_password,
            port=int(smb_port or 445),
            timeout_sec=int(smb_timeout_sec or 30),
            client_name=smb_client_name.strip() or None,
        )
        return AddressBookWorkbook(
            workbook=load_workbook(filename=data, read_only=True, data_only=True),
            source_label=raw_path,
        )

    local_path = Path(raw_path)
    if local_path.is_dir():
        local_path = local_path / "全站点地址.xlsx"
    if not local_path.exists():
        raise FileNotFoundError(f"地址簿不存在: {local_path}")
    return AddressBookWorkbook(
        workbook=load_workbook(filename=str(local_path), read_only=True, data_only=True),
        source_label=str(local_path),
    )
