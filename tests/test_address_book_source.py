import sys
import types
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook


def _workbook_bytes() -> bytes:
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "美国"
    ws.append(["收件人", "目的港"])
    ws.append(["Example Receiver", "ABE8"])
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()


class AddressBookSourceTests(unittest.TestCase):
    def test_local_path_loads_workbook_from_filesystem(self):
        import address_book_source

        path = Path("/tmp/address-book-source-test.xlsx")
        path.write_bytes(_workbook_bytes())
        try:
            result = address_book_source.open_address_book_workbook(
                path=str(path),
                smb_host="",
                smb_share="",
                smb_username="",
                smb_password="",
            )
            try:
                self.assertEqual(result.workbook.sheetnames, ["美国"])
                self.assertEqual(result.source_label, str(path))
            finally:
                result.close()
        finally:
            path.unlink(missing_ok=True)

    def test_smb_path_reads_workbook_via_smb_connection(self):
        captured = {}
        workbook_data = _workbook_bytes()

        class FakeSMBConnection:
            def __init__(self, username, password, client_machine_name, server_name, use_ntlm_v2, is_direct_tcp):
                captured["init"] = {
                    "username": username,
                    "password": password,
                    "client_machine_name": client_machine_name,
                    "server_name": server_name,
                    "use_ntlm_v2": use_ntlm_v2,
                    "is_direct_tcp": is_direct_tcp,
                }

            def connect(self, host, port, timeout):
                captured["connect"] = {"host": host, "port": port, "timeout": timeout}
                return True

            def retrieveFile(self, share, path, file_obj, timeout):
                captured["retrieve"] = {"share": share, "path": path, "timeout": timeout}
                file_obj.write(workbook_data)
                return object(), len(workbook_data)

            def close(self):
                captured["closed"] = True

        fake_smb_connection_module = types.ModuleType("smb.SMBConnection")
        fake_smb_connection_module.SMBConnection = FakeSMBConnection
        fake_smb_module = types.ModuleType("smb")

        with patch.dict(
            sys.modules,
            {
                "smb": fake_smb_module,
                "smb.SMBConnection": fake_smb_connection_module,
            },
        ):
            import address_book_source

            result = address_book_source.open_address_book_workbook(
                path="smb://192.168.0.45/供应链管理/2 物流发货管理/17.单证数据表维护/全站点地址.xlsx",
                smb_host="",
                smb_share="",
                smb_username="Logistics",
                smb_password="secret",
            )

        try:
            self.assertEqual(result.workbook.sheetnames, ["美国"])
            self.assertEqual(result.source_label, "smb://192.168.0.45/供应链管理/2 物流发货管理/17.单证数据表维护/全站点地址.xlsx")
            self.assertEqual(captured["init"]["username"], "Logistics")
            self.assertEqual(captured["init"]["password"], "secret")
            self.assertEqual(captured["connect"]["host"], "192.168.0.45")
            self.assertEqual(captured["connect"]["port"], 445)
            self.assertEqual(captured["retrieve"]["share"], "供应链管理")
            self.assertEqual(captured["retrieve"]["path"], "/2 物流发货管理/17.单证数据表维护/全站点地址.xlsx")
            self.assertTrue(captured["closed"])
        finally:
            result.close()


if __name__ == "__main__":
    unittest.main()
